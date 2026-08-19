from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo
import base64
import json
import html as html_lib
import os
import re
import secrets
import smtplib
import sqlite3
import tempfile
import threading
import uuid

import gspread
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    flash,
    redirect,
    render_template,
    request,
    stream_with_context,
    url_for,
)

try:
    from openai import OpenAI
except ImportError:
    OpenAI = None  # type: ignore[assignment]


BASE_DIR = Path(__file__).resolve().parent
RUNTIME_DATA_DIR = Path(tempfile.gettempdir()) if os.getenv("VERCEL") else BASE_DIR
DATABASE_PATH = RUNTIME_DATA_DIR / "portfolio.db"
load_dotenv(BASE_DIR / ".env.local")
load_dotenv(BASE_DIR / ".env")

import enquiry_store

app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.getenv("SECRET_KEY", "development-only-change-me"),
    MAX_CONTENT_LENGTH=12 * 1024 * 1024,
)

MAIL_USERNAME = os.getenv("MAIL_USERNAME", "").strip()
MAIL_PASSWORD = os.getenv("MAIL_PASSWORD", "").replace(" ", "").strip()
MAIL_RECEIVER = os.getenv(
    "MAIL_RECEIVER", "princechauhan31081997@gmail.com"
).strip()
ENQUIRY_EXCEL_PATH = RUNTIME_DATA_DIR / os.getenv(
    "ENQUIRY_EXCEL_FILE", "Portfolio_Enquiries.xlsx"
).strip()
EXCEL_LOCK = threading.Lock()
LOGO_PATH = BASE_DIR / "static" / "images" / "shubham-chauhan-logo.png"
INDIA_TIMEZONE = ZoneInfo("Asia/Kolkata")
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "").strip()
GOOGLE_SHEET_NAME = os.getenv("GOOGLE_SHEET_NAME", "Enquiries").strip()
GOOGLE_SERVICE_ACCOUNT_PATH = BASE_DIR / os.getenv(
    "GOOGLE_SERVICE_ACCOUNT_FILE", "google-service-account.json"
).strip()
GOOGLE_SERVICE_ACCOUNT_JSON_BASE64 = os.getenv(
    "GOOGLE_SERVICE_ACCOUNT_JSON_BASE64", ""
).strip()
GOOGLE_SHEET_LOCK = threading.Lock()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-5-mini").strip()

WHATSAPP_NUMBER = re.sub(
    r"\D",
    "",
    os.getenv("WHATSAPP_NUMBER", "918929932706"),
)

openai_client = (
    OpenAI(api_key=OPENAI_API_KEY)
    if OpenAI is not None and OPENAI_API_KEY
    else None
)


def database_connection() -> sqlite3.Connection:
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def initialise_database() -> None:
    with database_connection() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS contact_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT NOT NULL,
                phone TEXT,
                subject TEXT NOT NULL,
                message TEXT NOT NULL,
                email_status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL
            )
            """
        )
        connection.commit()


def safe_text(value: Any, maximum: int) -> str:
    return str(value or "").strip()[:maximum]


def looks_like_email(value: str) -> bool:
    return bool(
        re.fullmatch(
            r"[^@\s]+@[^@\s]+\.[^@\s]+",
            value,
        )
    )


def format_india_datetime(value: str | datetime | None = None) -> str:
    """Return a consistent, human-readable timestamp in Indian Standard Time."""
    if value is None:
        moment = datetime.now(timezone.utc)
    elif isinstance(value, datetime):
        moment = value
    else:
        moment = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if moment.tzinfo is None:
        moment = moment.replace(tzinfo=timezone.utc)
    return moment.astimezone(INDIA_TIMEZONE).strftime("%d %b %Y, %I:%M %p IST")


def parse_payment_date(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text or text.lower() in {"false", "none", "null"}:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("Payment Date must be YYYY-MM-DD or DD/MM/YYYY")


EXCEL_HEADERS = [
    "Enquiry ID",
    "Received At",
    "Client Name",
    "Email Address",
    "Phone / WhatsApp",
    "Project Subject",
    "Project Details",
    "Lead Status",
    "Email Delivery",
    "Source",
    "Enquiry Validity",
    "Validation Notes",
]

SYNC_HEADERS = [
    "Enquiry ID", "Received At", "Client Name", "Email Address",
    "Phone / WhatsApp", "Project Subject", "Project Details", "Project Amount",
    "Amount Received", "Payment Date", "Payment Status", "Lead Status",
    "Email Delivery", "Source", "Enquiry Validity", "Validation Notes",
    "Updated At", "Sync Status",
    "Quotation Number", "Quotation Amount", "Currency", "Delivery Days",
    "Valid Until", "Payment Terms", "Quotation Notes", "Quotation Status",
    "Quotation Sent At", "Send Thank You", "Send Quotation",
    "Payment Due Date", "Pending Amount", "Payment Email Status",
    "Payment Receipt Sent At", "Payment Reminder Sent At",
    "Send Payment Receipt", "Send Payment Reminder",
]

SYNC_COLUMN = {header: index + 1 for index, header in enumerate(SYNC_HEADERS)}


def quotation_values(enquiry_id: str, project_amount: Any = None) -> dict[str, Any]:
    """Build safe editable defaults for an initial project quotation."""
    amount = enquiry_store.parse_amount(project_amount)
    return {
        "quotation_number": f"QUO-{enquiry_id.removeprefix('ENQ-')}",
        "quotation_amount": amount,
        "quotation_currency": "INR",
        "delivery_days": 45,
        "quotation_valid_until": date.today() + timedelta(days=7),
        "payment_terms": "40% advance, 30% after milestone, 30% before final delivery",
        "quotation_notes": (
            "Preliminary quotation based on the submitted requirements. "
            "Final scope, timeline and price will be confirmed after discussion."
        ),
        "quotation_status": "Pending",
        "quotation_sent_at": None,
    }


def assess_enquiry_validity(email: str, phone: str, message: str) -> tuple[str, str]:
    """Classify basic lead quality without rejecting legitimate optional-phone leads."""
    notes: list[str] = []
    if not looks_like_email(email):
        return "Invalid", "Invalid email address"

    phone_digits = re.sub(r"\D", "", phone)
    if phone and not 7 <= len(phone_digits) <= 15:
        notes.append("Phone number needs review")
    if len(message.strip()) < 20:
        notes.append("Project details are very short")
    if len(re.findall(r"https?://|www\.", message, flags=re.IGNORECASE)) >= 3:
        notes.append("Multiple links detected")

    if any(note in notes for note in ("Phone number needs review", "Multiple links detected")):
        return "Review", "; ".join(notes)
    return "Valid", "; ".join(notes) or "Email and enquiry details passed validation"


def style_enquiry_sheet(sheet) -> None:
    """Apply a polished, reusable layout to the enquiry workbook."""
    navy = "172554"
    blue = "2563EB"
    white = "FFFFFF"
    muted = "64748B"
    border_colour = "CBD5E1"

    for merged_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged_range))
    sheet._images = []
    sheet.conditional_formatting._cf_rules.clear()
    sheet.data_validations.dataValidation = []

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.merge_cells("A1:B3")
    sheet.merge_cells("C1:L1")
    sheet["C1"] = "PORTFOLIO ENQUIRIES"
    sheet["C1"].font = Font(name="Aptos Display", size=20, bold=True, color=white)
    sheet["C1"].alignment = Alignment(horizontal="left", vertical="center")

    sheet.merge_cells("C2:L2")
    sheet["C2"] = "Professional lead register | Shubham Chauhan, Full Stack Developer"
    sheet["C2"].font = Font(name="Aptos", size=10, italic=True, color="CBD5E1")
    sheet["C2"].alignment = Alignment(vertical="center")
    for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=12):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=navy)
    sheet.row_dimensions[1].height = 38
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 8

    if LOGO_PATH.exists():
        logo = ExcelImage(LOGO_PATH)
        logo.width = 88
        logo.height = 88
        sheet.add_image(logo, "A1")

    thin_border = Border(bottom=Side(style="thin", color=border_colour))
    for column, header in enumerate(EXCEL_HEADERS, 1):
        cell = sheet.cell(row=4, column=column, value=header)
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
    sheet.row_dimensions[4].height = 28
    sheet.auto_filter.ref = "A4:L4"

    widths = {
        "A": 13, "B": 22, "C": 22, "D": 30, "E": 20,
        "F": 30, "G": 55, "H": 18, "I": 18, "J": 18,
        "K": 18, "L": 38,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    status_validation = DataValidation(
        type="list",
        formula1='"New,Contacted,Qualified,Proposal Sent,Won,Lost,Closed"',
        allow_blank=False,
    )
    status_validation.promptTitle = "Lead status"
    status_validation.prompt = "Select the current enquiry stage."
    status_validation.error = "Please select a status from the list."
    status_validation.errorTitle = "Invalid status"
    status_validation.showErrorMessage = True
    status_validation.showInputMessage = True
    sheet.add_data_validation(status_validation)
    status_validation.add("H5:H1048576")

    validity_colours = {"Valid": "DCFCE7", "Review": "FEF3C7", "Invalid": "FEE2E2"}
    for validity, colour in validity_colours.items():
        sheet.conditional_formatting.add(
            "K5:K1048576",
            FormulaRule(
                formula=[f'$K5="{validity}"'],
                fill=PatternFill("solid", fgColor=colour),
                font=Font(name="Aptos", bold=True, color=navy),
            ),
        )

    row_validity_colours = {"Valid": "F0FDF4", "Review": "FFFBEB", "Invalid": "FEF2F2"}
    for validity, colour in row_validity_colours.items():
        sheet.conditional_formatting.add(
            "A5:L1048576",
            FormulaRule(
                formula=[f'$K5="{validity}"'],
                fill=PatternFill("solid", fgColor=colour),
            ),
        )

    status_colours = {
        "New": "DBEAFE",
        "Contacted": "FEF3C7",
        "Qualified": "EDE9FE",
        "Proposal Sent": "FFEDD5",
        "Won": "DCFCE7",
        "Lost": "FEE2E2",
        "Closed": "E2E8F0",
    }
    for status, colour in status_colours.items():
        sheet.conditional_formatting.add(
            "H5:H1048576",
            FormulaRule(
                formula=[f'$H5="{status}"'],
                fill=PatternFill("solid", fgColor=colour),
                font=Font(name="Aptos", bold=True, color=navy),
            ),
        )


def append_enquiry_to_excel(
    *,
    enquiry_id: int,
    created_at: str,
    name: str,
    email: str,
    phone: str,
    subject: str,
    message: str,
    email_status: str,
) -> None:
    """Create the formatted workbook when needed and append one enquiry safely."""
    with EXCEL_LOCK:
        if ENQUIRY_EXCEL_PATH.exists():
            workbook = load_workbook(ENQUIRY_EXCEL_PATH)
            sheet = workbook["Enquiries"] if "Enquiries" in workbook.sheetnames else workbook.create_sheet("Enquiries")
            current_headers = [sheet.cell(4, column).value for column in range(1, len(EXCEL_HEADERS) + 1)]
            if current_headers != EXCEL_HEADERS:
                style_enquiry_sheet(sheet)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Enquiries"
            style_enquiry_sheet(sheet)

        row = max(sheet.max_row + 1, 5)
        validity, validation_notes = assess_enquiry_validity(email, phone, message)
        values = [
            f"ENQ-{enquiry_id:05d}",
            format_india_datetime(created_at),
            name,
            email,
            phone or "Not provided",
            subject,
            message,
            "New",
            "Sent" if email_status == "sent" else "Saved only",
            "Portfolio Website",
            validity,
            validation_notes,
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.font = Font(name="Aptos", size=10, color="1E293B")
            cell.alignment = Alignment(vertical="top", wrap_text=column in {6, 7, 12})
            cell.border = Border(bottom=Side(style="hair", color="E2E8F0"))
            if row % 2 == 0 and column != 8:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
        sheet.row_dimensions[row].height = 42
        sheet.auto_filter.ref = f"A4:L{row}"
        style_excel_lead_finder(workbook)
        workbook.save(ENQUIRY_EXCEL_PATH)


def style_excel_lead_finder(workbook) -> None:
    """Create a premium Enquiry ID lookup sheet for desktop Excel users."""
    if "Lead Finder" in workbook.sheetnames:
        del workbook["Lead Finder"]
    sheet = workbook.create_sheet("Lead Finder", 0)
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:L1")
    sheet["A1"] = "LEAD FINDER"
    sheet["A1"].font = Font(name="Aptos Display", size=20, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="172554")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 38
    sheet["A3"] = "Search by Enquiry ID"
    sheet["A3"].font = Font(name="Aptos", bold=True, color="172554")
    sheet["B3"] = ""
    sheet["B3"].fill = PatternFill("solid", fgColor="EFF6FF")
    sheet["B3"].border = Border(bottom=Side(style="medium", color="2563EB"))
    validation = DataValidation(type="list", formula1="'Enquiries'!$A$5:$A$1048576", allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(sheet["B3"])
    for column, header in enumerate(EXCEL_HEADERS, 1):
        cell = sheet.cell(5, column, header)
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    sheet["A6"] = '=IF(B3="","Select an Enquiry ID above",IFERROR(XLOOKUP(B3,Enquiries!A:A,Enquiries!A:L),"No enquiry found"))'
    sheet["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    for column, width in {"A":16,"B":23,"C":24,"D":30,"E":20,"F":30,"G":50,"H":18,"I":18,"J":20,"K":18,"L":38}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A5"


def append_enquiry_to_google_sheet(
    *,
    enquiry_id: int,
    created_at: str,
    name: str,
    email: str,
    phone: str,
    subject: str,
    message: str,
    email_status: str,
) -> bool:
    """Append an enquiry to Google Sheets immediately when credentials exist."""
    has_google_credentials = bool(GOOGLE_SERVICE_ACCOUNT_JSON_BASE64) or (
        GOOGLE_SERVICE_ACCOUNT_PATH.is_file()
    )
    if not GOOGLE_SHEET_ID or not has_google_credentials:
        app.logger.warning("Google Sheets credentials are missing; live sync skipped.")
        return False

    with GOOGLE_SHEET_LOCK:
        if GOOGLE_SERVICE_ACCOUNT_JSON_BASE64:
            credentials = json.loads(
                base64.b64decode(GOOGLE_SERVICE_ACCOUNT_JSON_BASE64).decode("utf-8")
            )
            client = gspread.service_account_from_dict(credentials)
        else:
            client = gspread.service_account(filename=str(GOOGLE_SERVICE_ACCOUNT_PATH))
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        try:
            sheet = spreadsheet.worksheet(GOOGLE_SHEET_NAME)
        except gspread.WorksheetNotFound:
            sheet = spreadsheet.add_worksheet(
                title=GOOGLE_SHEET_NAME,
                rows=1000,
                cols=len(EXCEL_HEADERS),
            )

        if sheet.row_values(1) != EXCEL_HEADERS:
            sheet.update(range_name="A1:L1", values=[EXCEL_HEADERS])
            sheet.freeze(rows=1)
            sheet.format(
                "A1:L1",
                {
                    "backgroundColor": {"red": 0.145, "green": 0.388, "blue": 0.922},
                    "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                    "horizontalAlignment": "LEFT",
                },
            )
            sheet.set_basic_filter("A1:L1000")
            spreadsheet.batch_update(
                {
                    "requests": [
                        {
                            "updateDimensionProperties": {
                                "range": {"sheetId": sheet.id, "dimension": "ROWS", "startIndex": 0, "endIndex": 1},
                                "properties": {"pixelSize": 42},
                                "fields": "pixelSize",
                            }
                        },
                        {
                            "updateDimensionProperties": {
                                "range": {"sheetId": sheet.id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 12},
                                "properties": {"pixelSize": 170},
                                "fields": "pixelSize",
                            }
                        },
                        {
                            "updateDimensionProperties": {
                                "range": {"sheetId": sheet.id, "dimension": "COLUMNS", "startIndex": 6, "endIndex": 7},
                                "properties": {"pixelSize": 360},
                                "fields": "pixelSize",
                            }
                        },
                        {
                            "repeatCell": {
                                "range": {"sheetId": sheet.id, "startRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 12},
                                "cell": {"userEnteredFormat": {"verticalAlignment": "TOP", "wrapStrategy": "WRAP", "padding": {"top": 8, "bottom": 8, "left": 8, "right": 8}}},
                                "fields": "userEnteredFormat(verticalAlignment,wrapStrategy,padding)",
                            }
                        },
                        {
                            "setDataValidation": {
                                "range": {"sheetId": sheet.id, "startRowIndex": 1, "startColumnIndex": 7, "endColumnIndex": 8},
                                "rule": {
                                    "condition": {"type": "ONE_OF_LIST", "values": [{"userEnteredValue": value} for value in ["New", "Contacted", "Qualified", "Proposal Sent", "Won", "Lost", "Closed"]]},
                                    "strict": True,
                                    "showCustomUi": True,
                                },
                            }
                        },
                        *[
                            {
                                "addConditionalFormatRule": {
                                    "rule": {
                                        "ranges": [{"sheetId": sheet.id, "startRowIndex": 1, "startColumnIndex": 7, "endColumnIndex": 8}],
                                        "booleanRule": {
                                            "condition": {"type": "TEXT_EQ", "values": [{"userEnteredValue": status}]},
                                            "format": {"backgroundColor": colour, "textFormat": {"bold": True}},
                                        },
                                    },
                                    "index": 0,
                                }
                            }
                            for status, colour in {
                                "New": {"red": .86, "green": .92, "blue": .99},
                                "Contacted": {"red": .99, "green": .94, "blue": .76},
                                "Qualified": {"red": .92, "green": .89, "blue": .99},
                                "Proposal Sent": {"red": 1, "green": .91, "blue": .82},
                                "Won": {"red": .86, "green": .97, "blue": .89},
                                "Lost": {"red": .99, "green": .88, "blue": .88},
                                "Closed": {"red": .89, "green": .91, "blue": .94},
                            }.items()
                        ],
                        *[
                            {
                                "addConditionalFormatRule": {
                                    "rule": {
                                        "ranges": [{"sheetId": sheet.id, "startRowIndex": 1, "startColumnIndex": 10, "endColumnIndex": 11}],
                                        "booleanRule": {
                                            "condition": {"type": "TEXT_EQ", "values": [{"userEnteredValue": validity}]},
                                            "format": {"backgroundColor": colour, "textFormat": {"bold": True}},
                                        },
                                    },
                                    "index": 0,
                                }
                            }
                            for validity, colour in {
                                "Valid": {"red": .86, "green": .97, "blue": .89},
                                "Review": {"red": .99, "green": .94, "blue": .76},
                                "Invalid": {"red": .99, "green": .88, "blue": .88},
                            }.items()
                        ],
                        *[
                            {
                                "addConditionalFormatRule": {
                                    "rule": {
                                        "ranges": [{"sheetId": sheet.id, "startRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 12}],
                                        "booleanRule": {
                                            "condition": {"type": "CUSTOM_FORMULA", "values": [{"userEnteredValue": f'=$K2="{validity}"'}]},
                                            "format": {"backgroundColor": colour},
                                        },
                                    },
                                    "index": 0,
                                }
                            }
                            for validity, colour in {
                                "Valid": {"red": .94, "green": .99, "blue": .96},
                                "Review": {"red": 1, "green": .98, "blue": .92},
                                "Invalid": {"red": 1, "green": .95, "blue": .95},
                            }.items()
                        ],
                    ]
                }
            )

        validity, validation_notes = assess_enquiry_validity(email, phone, message)
        sheet.append_row(
            [
                f"ENQ-{enquiry_id:05d}",
                format_india_datetime(created_at),
                name,
                email,
                phone or "Not provided",
                subject,
                message,
                "New",
                "Sent" if email_status == "sent" else "Saved only",
                "Portfolio Website",
                validity,
                validation_notes,
            ],
            value_input_option="USER_ENTERED",
            table_range="A:L",
        )
        return True


def google_sheet_client():
    if GOOGLE_SERVICE_ACCOUNT_JSON_BASE64:
        credentials = json.loads(
            base64.b64decode(GOOGLE_SERVICE_ACCOUNT_JSON_BASE64).decode("utf-8")
        )
        return gspread.service_account_from_dict(credentials)
    if GOOGLE_SERVICE_ACCOUNT_PATH.is_file():
        return gspread.service_account(filename=str(GOOGLE_SERVICE_ACCOUNT_PATH))
    raise RuntimeError("Google service-account credentials are not configured")


def synced_google_worksheet():
    if not GOOGLE_SHEET_ID:
        raise RuntimeError("GOOGLE_SHEET_ID is not configured")
    spreadsheet = google_sheet_client().open_by_key(GOOGLE_SHEET_ID)
    try:
        sheet = spreadsheet.worksheet(GOOGLE_SHEET_NAME)
    except gspread.WorksheetNotFound:
        sheet = spreadsheet.add_worksheet(
            title=GOOGLE_SHEET_NAME, rows=1000, cols=len(SYNC_HEADERS)
        )
    current_headers = sheet.row_values(1)
    if current_headers == EXCEL_HEADERS:
        sheet.insert_cols([["Project Amount"]], col=8)
        sheet.resize(cols=len(SYNC_HEADERS))
        current_headers = sheet.row_values(1)
    if (
        "Send Thank You" not in current_headers
        and current_headers
        and current_headers[-1] == "Send Quotation"
    ):
        sheet.insert_cols([["Send Thank You"]], col=SYNC_COLUMN["Send Thank You"])
        sheet.resize(cols=len(SYNC_HEADERS))
        current_headers = sheet.row_values(1)
    if current_headers != SYNC_HEADERS:
        sheet.resize(cols=len(SYNC_HEADERS))
        sheet.update(range_name="A1:AJ1", values=[SYNC_HEADERS])
        sheet.freeze(rows=1)
        sheet.format(
            "A1:R1",
            {
                "backgroundColor": {"red": 0.09, "green": 0.15, "blue": 0.33},
                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            },
        )
        sheet.set_basic_filter("A1:AJ1000")
        style_synced_enquiry_sheet(spreadsheet, sheet)
    spreadsheet.batch_update({"requests": [{
        "setDataValidation": {
            "range": {
                "sheetId": sheet.id, "startRowIndex": 1,
                "endRowIndex": sheet.row_count,
                "startColumnIndex": SYNC_COLUMN["Send Thank You"] - 1,
                "endColumnIndex": SYNC_COLUMN["Send Quotation"],
            },
            "rule": {
                "condition": {"type": "BOOLEAN"},
                "strict": True, "showCustomUi": True,
            },
        }
    }, {
        "setDataValidation": {
            "range": {
                "sheetId": sheet.id, "startRowIndex": 1,
                "endRowIndex": sheet.row_count,
                "startColumnIndex": SYNC_COLUMN["Send Payment Receipt"] - 1,
                "endColumnIndex": SYNC_COLUMN["Send Payment Reminder"],
            },
            "rule": {
                "condition": {"type": "BOOLEAN"},
                "strict": True, "showCustomUi": True,
            },
        }
    }]})
    configure_google_sheet_dashboard(spreadsheet, sheet)
    configure_requirements_sheet(spreadsheet)
    return sheet


def configure_requirements_sheet(spreadsheet) -> None:
    """Maintain a polished requirements view backed by the Enquiries sheet."""
    try:
        requirements = spreadsheet.worksheet("Requirements")
    except gspread.WorksheetNotFound:
        requirements = spreadsheet.add_worksheet(title="Requirements", rows=1000, cols=15)
    requirements.resize(rows=max(requirements.row_count, 1000), cols=15)
    formula = (
        '=QUERY(Enquiries!A:Z,"select A,B,C,D,E,F,G,S,T,U,V,W,X,Y,Z '
        'where A is not null",1)'
    )
    current_formula = requirements.acell("A1", value_render_option="FORMULA").value
    if current_formula != formula:
        requirements.clear()
        requirements.update_acell("A1", formula)
    requirements.freeze(rows=1, cols=2)
    requirements.format("A1:O1", {
        "backgroundColor": {"red": 0.09, "green": 0.15, "blue": 0.33},
        "textFormat": {
            "bold": True, "fontSize": 10,
            "foregroundColor": {"red": 1, "green": 1, "blue": 1},
        },
        "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE",
        "wrapStrategy": "WRAP",
    })
    requirements.format("A2:O1000", {
        "verticalAlignment": "MIDDLE", "wrapStrategy": "WRAP",
        "textFormat": {"fontSize": 10},
    })
    widths = [155, 155, 175, 220, 150, 220, 380, 165, 140, 90, 115, 130, 290, 310, 140]
    spreadsheet.batch_update({"requests": [{
        "updateDimensionProperties": {
            "range": {
                "sheetId": requirements.id, "dimension": "COLUMNS",
                "startIndex": index, "endIndex": index + 1,
            },
            "properties": {"pixelSize": width}, "fields": "pixelSize",
        }
    } for index, width in enumerate(widths)]})


def style_synced_enquiry_sheet(spreadsheet, sheet) -> None:
    """Apply a compact CRM-style layout without changing enquiry data."""
    sheet.freeze(rows=1, cols=2)
    sheet.format("A1:AJ1", {
        "backgroundColor": {"red": 0.055, "green": 0.102, "blue": 0.216},
        "textFormat": {
            "bold": True,
            "fontSize": 10,
            "foregroundColor": {"red": 1, "green": 1, "blue": 1},
        },
        "horizontalAlignment": "CENTER",
        "verticalAlignment": "MIDDLE",
        "wrapStrategy": "WRAP",
    })
    sheet.format("A2:AJ1000", {
        "backgroundColor": {"red": 1, "green": 1, "blue": 1},
        "verticalAlignment": "MIDDLE",
        "wrapStrategy": "WRAP",
        "textFormat": {"fontSize": 10, "foregroundColor": {
            "red": 0.118, "green": 0.161, "blue": 0.231,
        }},
    })
    # Finance and quotation blocks use distinct header colours for quick scanning.
    sheet.format("H1:K1", {"backgroundColor": {
        "red": 0.11, "green": 0.35, "blue": 0.75,
    }})
    sheet.format("S1:AC1", {"backgroundColor": {
        "red": 0.145, "green": 0.388, "blue": 0.922,
    }})
    sheet.format("AD1:AJ1", {"backgroundColor": {
        "red": 0.035, "green": 0.22, "blue": 0.52,
    }})
    for editable_range in ("J2:J1000", "W2:W1000", "AD2:AD1000", "AB2:AC1000", "AI2:AJ1000"):
        sheet.format(editable_range, {
            "backgroundColor": {"red": 0.937, "green": 0.965, "blue": 1},
            "textFormat": {"foregroundColor": {"red": 0.055, "green": 0.102, "blue": 0.216}},
        })
    widths = [
        155, 155, 175, 220, 150, 220, 380, 130, 130, 125, 130, 140,
        125, 150, 130, 280, 165, 125, 165, 140, 90, 115, 130, 290, 310,
        140, 170, 135, 135, 135, 145, 170, 175, 185, 155, 165,
    ]
    requests = [{
        "updateDimensionProperties": {
            "range": {
                "sheetId": sheet.id, "dimension": "ROWS",
                "startIndex": 0, "endIndex": 1,
            },
            "properties": {"pixelSize": 52}, "fields": "pixelSize",
        }
    }, {
        "updateDimensionProperties": {
            "range": {
                "sheetId": sheet.id, "dimension": "ROWS",
                "startIndex": 1, "endIndex": min(sheet.row_count, 1000),
            },
            "properties": {"pixelSize": 44}, "fields": "pixelSize",
        }
    }]
    requests.extend({
        "updateDimensionProperties": {
            "range": {
                "sheetId": sheet.id, "dimension": "COLUMNS",
                "startIndex": index, "endIndex": index + 1,
            },
            "properties": {"pixelSize": width}, "fields": "pixelSize",
        }
    } for index, width in enumerate(widths))
    spreadsheet.batch_update({"requests": requests})


def configure_google_sheet_dashboard(spreadsheet, enquiry_sheet) -> None:
    """Keep amount/status rules and the searchable all-leads view schema-safe."""
    spreadsheet.batch_update(
        {
            "requests": [
                *[
                    {
                        "setDataValidation": {
                            "range": {
                                "sheetId": enquiry_sheet.id,
                                "startRowIndex": 1,
                                "startColumnIndex": SYNC_COLUMN[header] - 1,
                                "endColumnIndex": SYNC_COLUMN[header],
                            },
                            "rule": {
                                "condition": {"type": "DATE_IS_VALID"},
                                "strict": True,
                                "showCustomUi": True,
                            },
                        }
                    }
                    for header in ("Payment Date", "Valid Until", "Payment Due Date")
                ],
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": enquiry_sheet.id,
                            "startRowIndex": 1,
                            "startColumnIndex": SYNC_COLUMN["Payment Due Date"] - 1,
                            "endColumnIndex": SYNC_COLUMN["Payment Due Date"],
                        },
                        "cell": {"userEnteredFormat": {"numberFormat": {"type": "DATE", "pattern": "dd mmm yyyy"}}},
                        "fields": "userEnteredFormat.numberFormat",
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": enquiry_sheet.id,
                            "startRowIndex": 1,
                            "startColumnIndex": SYNC_COLUMN["Valid Until"] - 1,
                            "endColumnIndex": SYNC_COLUMN["Valid Until"],
                        },
                        "cell": {"userEnteredFormat": {"numberFormat": {"type": "DATE", "pattern": "dd mmm yyyy"}}},
                        "fields": "userEnteredFormat.numberFormat",
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": enquiry_sheet.id,
                            "startRowIndex": 1,
                            "startColumnIndex": SYNC_COLUMN["Pending Amount"] - 1,
                            "endColumnIndex": SYNC_COLUMN["Pending Amount"],
                        },
                        "cell": {"userEnteredFormat": {"numberFormat": {"type": "CURRENCY", "pattern": "₹#,##0.00"}}},
                        "fields": "userEnteredFormat.numberFormat",
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": enquiry_sheet.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 9,
                            "endColumnIndex": 10,
                        },
                        "cell": {"userEnteredFormat": {"numberFormat": {"type": "DATE", "pattern": "dd mmm yyyy"}}},
                        "fields": "userEnteredFormat.numberFormat",
                    }
                },
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": enquiry_sheet.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 7,
                            "endColumnIndex": 9,
                        },
                        "rule": None,
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": enquiry_sheet.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 7,
                            "endColumnIndex": 9,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "numberFormat": {"type": "CURRENCY", "pattern": "₹#,##0.00"}
                            }
                        },
                        "fields": "userEnteredFormat.numberFormat",
                    }
                },
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": enquiry_sheet.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 10,
                            "endColumnIndex": 11,
                        },
                        "rule": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": value}
                                    for value in ["Pending", "Partial", "Paid", "Refunded"]
                                ],
                            },
                            "strict": True,
                            "showCustomUi": True,
                        },
                    }
                },
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": enquiry_sheet.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 11,
                            "endColumnIndex": 12,
                        },
                        "rule": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": value}
                                    for value in [
                                        "New", "Contacted", "Qualified", "Proposal Sent",
                                        "Won", "Lost", "Closed",
                                    ]
                                ],
                            },
                            "strict": True,
                            "showCustomUi": True,
                        },
                    }
                },
            ]
        }
    )
    try:
        finder = spreadsheet.worksheet("Lead Finder")
    except gspread.WorksheetNotFound:
        finder = spreadsheet.add_worksheet(title="Lead Finder", rows=500, cols=23)
    finder.resize(rows=max(finder.row_count, 500), cols=len(SYNC_HEADERS))
    metadata = spreadsheet.fetch_sheet_metadata()
    finder_metadata = next(
        item for item in metadata["sheets"]
        if item["properties"]["sheetId"] == finder.id
    )
    summary_unmerge_requests = []
    for merged_range in finder_metadata.get("merges", []):
        if (
            merged_range.get("startRowIndex", 0) < 6
            and merged_range.get("endRowIndex", 0) > 4
            and merged_range.get("startColumnIndex", 0) < 15
            and merged_range.get("endColumnIndex", 0) > 4
        ):
            summary_unmerge_requests.append({"unmergeCells": {"range": merged_range}})
    if summary_unmerge_requests:
        spreadsheet.batch_update({"requests": summary_unmerge_requests})

    finder.update(
        range_name="E5:O6",
        values=[
            [
                "WON VALUE", "LOST VALUE", "ACTIVE PIPELINE", "TOTAL QUOTED",
                "WON PROJECTS", "LOST PROJECTS", "ACTIVE PROJECTS", "", "", "", "",
            ],
            [
                '=SUMIFS(Enquiries!H2:H,Enquiries!I2:I,"Won",Enquiries!L2:L,"Valid")',
                '=SUMIFS(Enquiries!H2:H,Enquiries!I2:I,"Lost",Enquiries!L2:L,"Valid")',
                '=SUMIFS(Enquiries!H2:H,Enquiries!L2:L,"Valid",Enquiries!I2:I,"<>Won",Enquiries!I2:I,"<>Lost",Enquiries!I2:I,"<>Closed")',
                '=SUMIFS(Enquiries!H2:H,Enquiries!L2:L,"Valid")',
                '=COUNTIFS(Enquiries!I2:I,"Won",Enquiries!L2:L,"Valid")',
                '=COUNTIFS(Enquiries!I2:I,"Lost",Enquiries!L2:L,"Valid")',
                '=COUNTIFS(Enquiries!A2:A,"<>",Enquiries!L2:L,"Valid",Enquiries!I2:I,"<>Won",Enquiries!I2:I,"<>Lost",Enquiries!I2:I,"<>Closed")',
                "", "", "", "",
            ],
        ],
        value_input_option="USER_ENTERED",
    )
    finder.format(
        "E5:O6",
        {
            "backgroundColor": {"red": 0.945, "green": 0.961, "blue": 0.988},
            "textFormat": {"bold": True, "foregroundColor": {"red": 0.09, "green": 0.15, "blue": 0.33}},
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
        },
    )
    finder.format(
        "E6:H6",
        {
            "numberFormat": {"type": "CURRENCY", "pattern": "₹#,##0.00"},
            "textFormat": {"bold": True, "fontSize": 12},
        },
    )
    finder.format(
        "E5:O5",
        {
            "backgroundColor": {"red": 0.09, "green": 0.15, "blue": 0.33},
            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        },
    )
    finder.batch_clear(["E5:O6"])
    finder.format(
        "E5:O6",
        {
            "backgroundColor": {"red": 1, "green": 1, "blue": 1},
            "textFormat": {"bold": False, "foregroundColor": {"red": 0.12, "green": 0.12, "blue": 0.12}},
        },
    )

    try:
        summary = spreadsheet.worksheet("Business Summary")
    except gspread.WorksheetNotFound:
        summary = spreadsheet.add_worksheet(title="Business Summary", rows=60, cols=12)
    summary.resize(rows=max(summary.row_count, 60), cols=max(summary.col_count, 20))
    spreadsheet.batch_update(
        {
            "requests": [
                {
                    "updateSheetProperties": {
                        "properties": {"sheetId": summary.id, "index": 0},
                        "fields": "index",
                    }
                }
            ]
        }
    )
    summary.batch_clear(["A1:L60"])
    summary.update(
        range_name="A1:H13",
        values=[
            ["BUSINESS SUMMARY", "", "", "", "", "", "", ""],
            ["Live revenue and project performance — valid enquiries only", "", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", ""],
            ["TOTAL RECEIVED", "", "OUTSTANDING", "", "ACTIVE PIPELINE", "", "TOTAL QUOTED", ""],
            ['=SUMIFS(Enquiries!I2:I,Enquiries!O2:O,"Valid")', "", '=SUMIFS(Enquiries!H2:H,Enquiries!O2:O,"Valid")-SUMIFS(Enquiries!I2:I,Enquiries!O2:O,"Valid")', "", '=SUMIFS(Enquiries!H2:H,Enquiries!O2:O,"Valid",Enquiries!L2:L,"<>Won",Enquiries!L2:L,"<>Lost",Enquiries!L2:L,"<>Closed")', "", '=SUMIFS(Enquiries!H2:H,Enquiries!O2:O,"Valid")', ""],
            ["", "", "", "", "", "", "", ""],
            ["PROJECT PERFORMANCE", "", "", "", "", "", "", ""],
            ["WON PROJECTS", "", "LOST PROJECTS", "", "ACTIVE PROJECTS", "", "TOTAL VALID LEADS", ""],
            ['=COUNTIFS(Enquiries!L2:L,"Won",Enquiries!O2:O,"Valid")', "", '=COUNTIFS(Enquiries!L2:L,"Lost",Enquiries!O2:O,"Valid")', "", '=COUNTIFS(Enquiries!A2:A,"<>",Enquiries!O2:O,"Valid",Enquiries!L2:L,"<>Won",Enquiries!L2:L,"<>Lost",Enquiries!L2:L,"<>Closed")', "", '=COUNTIFS(Enquiries!A2:A,"<>",Enquiries!O2:O,"Valid")', ""],
            ["", "", "", "", "", "", "", ""],
            ["CALCULATION RULES", "", "", "", "", "", "", ""],
            ["Total Received uses Amount Received. Monthly income uses Payment Date for accurate cash-flow reporting.", "", "", "", "", "", "", ""],
            ["Outstanding = Total Quoted − Total Received. All cards and charts update automatically.", "", "", "", "", "", "", ""],
        ],
        value_input_option="USER_ENTERED",
    )
    summary.freeze(rows=2)
    summary.format(
        "A1:H2",
        {
            "backgroundColor": {"red": 0.04, "green": 0.09, "blue": 0.22},
            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        },
    )
    summary.format("A1", {"textFormat": {"bold": True, "fontSize": 20, "foregroundColor": {"red": 1, "green": 1, "blue": 1}}})
    for card_range in ["A4:B5", "C4:D5", "E4:F5", "G4:H5", "A8:B9", "C8:D9", "E8:F9", "G8:H9"]:
        summary.format(
            card_range,
            {
                "backgroundColor": {"red": 0.94, "green": 0.96, "blue": 0.99},
                "textFormat": {"bold": True, "foregroundColor": {"red": 0.07, "green": 0.16, "blue": 0.34}},
                "verticalAlignment": "MIDDLE",
            },
        )
    summary.format("A5:H5", {"numberFormat": {"type": "CURRENCY", "pattern": "₹#,##0.00"}, "textFormat": {"bold": True, "fontSize": 16}})
    summary.format("A9:H9", {"textFormat": {"bold": True, "fontSize": 16}})
    summary.format("A7:H7", {"backgroundColor": {"red": 0.145, "green": 0.388, "blue": 0.922}, "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}}})
    summary.format("A11:H13", {"backgroundColor": {"red": 0.97, "green": 0.98, "blue": 1}, "textFormat": {"foregroundColor": {"red": 0.25, "green": 0.3, "blue": 0.4}}})
    summary.update(
        range_name="J2:N8",
        values=[
            ["STATUS", "PROJECTS", "", "METRIC", "VALUE"],
            ["New", '=COUNTIFS(Enquiries!L2:L,"New",Enquiries!O2:O,"Valid")', "", "Received", '=SUMIFS(Enquiries!I2:I,Enquiries!O2:O,"Valid")'],
            ["Contacted", '=COUNTIFS(Enquiries!L2:L,"Contacted",Enquiries!O2:O,"Valid")', "", "Outstanding", '=SUMIFS(Enquiries!H2:H,Enquiries!O2:O,"Valid")-SUMIFS(Enquiries!I2:I,Enquiries!O2:O,"Valid")'],
            ["Qualified", '=COUNTIFS(Enquiries!L2:L,"Qualified",Enquiries!O2:O,"Valid")', "", "Active", '=SUMIFS(Enquiries!H2:H,Enquiries!O2:O,"Valid",Enquiries!L2:L,"<>Won",Enquiries!L2:L,"<>Lost",Enquiries!L2:L,"<>Closed")'],
            ["Proposal Sent", '=COUNTIFS(Enquiries!L2:L,"Proposal Sent",Enquiries!O2:O,"Valid")', "", "Total Quoted", '=SUMIFS(Enquiries!H2:H,Enquiries!O2:O,"Valid")'],
            ["Won", '=COUNTIFS(Enquiries!L2:L,"Won",Enquiries!O2:O,"Valid")', "", "", ""],
            ["Lost", '=COUNTIFS(Enquiries!L2:L,"Lost",Enquiries!O2:O,"Valid")', "", "", ""],
        ],
        value_input_option="USER_ENTERED",
    )
    summary.update(range_name="J9:K9", values=[["Closed", '=COUNTIFS(Enquiries!L2:L,"Closed",Enquiries!O2:O,"Valid")']], value_input_option="USER_ENTERED")
    summary.update_acell(
        "P2",
        '=IFERROR(QUERY({ARRAYFORMULA(IF(Enquiries!J2:J="","",TEXT(Enquiries!J2:J,"MMM YYYY"))),Enquiries!I2:I,Enquiries!O2:O},"select Col1,sum(Col2) where Col1 is not null and Col3 = \'Valid\' group by Col1 label Col1 \'Month\',sum(Col2) \'Income\'",0),{"Month","Income"})',
    )
    summary.update_acell(
        "S2",
        '=IFERROR(QUERY({Enquiries!F2:F,Enquiries!I2:I,Enquiries!O2:O},"select Col1,sum(Col2) where Col1 is not null and Col3 = \'Valid\' group by Col1 order by sum(Col2) desc limit 10 label Col1 \'Project\',sum(Col2) \'Received\'",0),{"Project","Received"})',
    )

    chart_metadata = spreadsheet.fetch_sheet_metadata()
    summary_metadata = next(
        item for item in chart_metadata["sheets"]
        if item["properties"]["sheetId"] == summary.id
    )
    chart_requests = [
        {"deleteEmbeddedObject": {"objectId": chart["chartId"]}}
        for chart in summary_metadata.get("charts", [])
    ]
    chart_requests.extend(
        [
            {
                "addChart": {
                    "chart": {
                        "spec": {
                            "title": "Project Status Distribution",
                            "subtitle": "Valid enquiries by current lead status",
                            "hiddenDimensionStrategy": "SHOW_ALL",
                            "pieChart": {
                                "legendPosition": "RIGHT_LEGEND",
                                "pieHole": 0.48,
                                "domain": {"sourceRange": {"sources": [{"sheetId": summary.id, "startRowIndex": 2, "endRowIndex": 9, "startColumnIndex": 9, "endColumnIndex": 10}]}},
                                "series": {"sourceRange": {"sources": [{"sheetId": summary.id, "startRowIndex": 2, "endRowIndex": 9, "startColumnIndex": 10, "endColumnIndex": 11}]}},
                            },
                        },
                        "position": {"overlayPosition": {"anchorCell": {"sheetId": summary.id, "rowIndex": 14, "columnIndex": 0}, "widthPixels": 570, "heightPixels": 360}},
                    }
                }
            },
            {
                "addChart": {
                    "chart": {
                        "spec": {
                            "title": "Monthly Income Trend",
                            "subtitle": "Actual payments grouped by Payment Date",
                            "hiddenDimensionStrategy": "SHOW_ALL",
                            "basicChart": {
                                "chartType": "LINE",
                                "legendPosition": "NO_LEGEND",
                                "headerCount": 0,
                                "axis": [
                                    {"position": "BOTTOM_AXIS", "title": "Month"},
                                    {"position": "LEFT_AXIS", "title": "Amount (₹)"},
                                ],
                                "domains": [{"domain": {"sourceRange": {"sources": [{"sheetId": summary.id, "startRowIndex": 2, "endRowIndex": 50, "startColumnIndex": 15, "endColumnIndex": 16}]}}}],
                                "series": [{"series": {"sourceRange": {"sources": [{"sheetId": summary.id, "startRowIndex": 2, "endRowIndex": 50, "startColumnIndex": 16, "endColumnIndex": 17}]}}, "targetAxis": "LEFT_AXIS"}],
                            },
                        },
                        "position": {"overlayPosition": {"anchorCell": {"sheetId": summary.id, "rowIndex": 14, "columnIndex": 6}, "widthPixels": 570, "heightPixels": 360}},
                    }
                }
            },
            {
                "addChart": {
                    "chart": {
                        "spec": {
                            "title": "Top Projects by Amount Received",
                            "subtitle": "Up to 10 highest-earning projects",
                            "hiddenDimensionStrategy": "SHOW_ALL",
                            "basicChart": {
                                "chartType": "BAR",
                                "legendPosition": "NO_LEGEND",
                                "headerCount": 0,
                                "axis": [
                                    {"position": "LEFT_AXIS", "title": "Project"},
                                    {"position": "BOTTOM_AXIS", "title": "Amount Received (₹)"},
                                ],
                                "domains": [{"domain": {"sourceRange": {"sources": [{"sheetId": summary.id, "startRowIndex": 2, "endRowIndex": 13, "startColumnIndex": 18, "endColumnIndex": 19}]}}}],
                                "series": [{"series": {"sourceRange": {"sources": [{"sheetId": summary.id, "startRowIndex": 2, "endRowIndex": 13, "startColumnIndex": 19, "endColumnIndex": 20}]}}, "targetAxis": "BOTTOM_AXIS"}],
                            },
                        },
                        "position": {"overlayPosition": {"anchorCell": {"sheetId": summary.id, "rowIndex": 34, "columnIndex": 0}, "widthPixels": 1140, "heightPixels": 420}},
                    }
                }
            },
        ]
    )
    chart_requests.append(
        {
            "updateDimensionProperties": {
                "range": {"sheetId": summary.id, "dimension": "COLUMNS", "startIndex": 9, "endIndex": 20},
                "properties": {"hiddenByUser": True},
                "fields": "hiddenByUser",
            }
        }
    )
    spreadsheet.batch_update({"requests": chart_requests})

    try:
        finder = spreadsheet.worksheet("Lead Finder")
    except gspread.WorksheetNotFound:
        finder = spreadsheet.add_worksheet(title="Lead Finder", rows=500, cols=23)
    finder.resize(rows=max(finder.row_count, 500), cols=max(finder.col_count, 23))
    desired_formula = (
        '=IF(B5="",IFERROR(FILTER(Enquiries!A2:AJ,Enquiries!A2:A<>""),"No enquiries yet"),'
        'IFERROR(FILTER(Enquiries!A2:AJ,REGEXMATCH(LOWER(Enquiries!A2:A&" "&'
        'Enquiries!C2:C&" "&Enquiries!D2:D&" "&Enquiries!F2:F),LOWER(B5))),'
        '"No matching enquiries"))'
    )
    current_header = finder.get("A8:AJ8")
    current_formula = finder.acell("A9", value_render_option="FORMULA").value
    if current_header != [SYNC_HEADERS] or current_formula != desired_formula:
        finder.batch_clear(["A8:AJ500"])
        finder.update(
            range_name="A3:O3",
            values=[["All enquiries are shown below. Type an ID, client name, email or project subject in B5 to search."] + [""] * 14],
        )
        finder.update(range_name="A5:B5", values=[["SEARCH", ""]])
        finder.update(range_name="A8:AJ8", values=[SYNC_HEADERS])
        finder.update_acell("A9", desired_formula)
        finder.freeze(rows=8)
        finder.format(
            "A8:AJ8",
            {
                "backgroundColor": {"red": 0.145, "green": 0.388, "blue": 0.922},
                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            },
        )
    spreadsheet.batch_update(
        {
            "requests": [
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": finder.id,
                            "startRowIndex": 4,
                            "endRowIndex": 5,
                            "startColumnIndex": 1,
                            "endColumnIndex": 2,
                        },
                        "rule": None,
                    }
                }
            ]
        }
    )


def upsert_enquiry_to_google_sheet(record: dict[str, Any]) -> int:
    """Idempotently update or insert one enquiry using its permanent ID."""
    with GOOGLE_SHEET_LOCK:
        sheet = synced_google_worksheet()
        enquiry_id = str(record["enquiry_id"])
        ids = sheet.col_values(1)
        row_number = ids.index(enquiry_id) + 1 if enquiry_id in ids else len(ids) + 1
        amount = record.get("project_amount")
        amount_received = record.get("amount_received")
        pending_amount = max((amount or 0) - (amount_received or 0), 0) if amount is not None else ""
        values = [[
            enquiry_id,
            format_india_datetime(record.get("received_at")),
            record.get("client_name", ""), record.get("email", ""),
            record.get("phone", "") or "Not provided",
            record.get("project_subject", ""), record.get("project_details", ""),
            float(amount) if amount is not None else "",
            float(amount_received) if amount_received is not None else 0,
            str(record.get("payment_date") or ""), record.get("payment_status", "Pending"),
            record.get("lead_status", "New"), record.get("email_status", "Pending"),
            record.get("source", "Portfolio Website"), record.get("validity", "Valid"),
            record.get("validation_notes", ""), format_india_datetime(), "Synced",
            record.get("quotation_number", ""),
            float(record["quotation_amount"]) if record.get("quotation_amount") is not None else "",
            record.get("quotation_currency", "INR"), record.get("delivery_days", ""),
            str(record.get("quotation_valid_until") or ""), record.get("payment_terms", ""),
            record.get("quotation_notes", ""), record.get("quotation_status", "Pending"),
            format_india_datetime(record.get("quotation_sent_at")) if record.get("quotation_sent_at") else "",
            "", "",
            str(record.get("payment_due_date") or ""),
            float(pending_amount) if pending_amount != "" else "",
            record.get("payment_email_status", "Pending"),
            format_india_datetime(record.get("payment_receipt_sent_at")) if record.get("payment_receipt_sent_at") else "",
            format_india_datetime(record.get("payment_reminder_sent_at")) if record.get("payment_reminder_sent_at") else "",
            "", "",
        ]]
        sheet.update(range_name=f"A{row_number}:AJ{row_number}", values=values)
        enquiry_store.mark_synced(enquiry_id, row_number)
        return row_number


def sync_google_sheet_to_database() -> dict[str, int]:
    """Upsert manual Sheet rows into Postgres; IDs prevent duplicate records."""
    sheet = synced_google_worksheet()
    rows = sheet.get_all_records(expected_headers=SYNC_HEADERS)
    result = {"created_or_updated": 0, "skipped": 0, "ids_assigned": 0}
    for index, row in enumerate(rows, start=2):
        name = safe_text(row.get("Client Name"), 100)
        email = safe_text(row.get("Email Address"), 180)
        subject = safe_text(row.get("Project Subject"), 160)
        details = safe_text(row.get("Project Details"), 5000)
        # Formatting or stale status text can make an otherwise empty row appear
        # in get_all_records(); silently clear it instead of displaying an error.
        if not any([name, email, subject, details, safe_text(row.get("Phone / WhatsApp"), 40)]):
            if row.get("Sync Status"):
                sheet.update_cell(index, SYNC_COLUMN["Sync Status"], "")
            continue
        if not all([name, email, subject, details]) or not looks_like_email(email):
            sheet.update_cell(index, SYNC_COLUMN["Sync Status"], "Error: required fields or email invalid")
            result["skipped"] += 1
            continue
        enquiry_id = safe_text(row.get("Enquiry ID"), 40)
        if not enquiry_id:
            enquiry_id = f"ENQ-{uuid.uuid4().hex[:12].upper()}"
            sheet.update_cell(index, 1, enquiry_id)
            result["ids_assigned"] += 1
        existing_record = enquiry_store.get_enquiry(enquiry_id)
        validity, notes = assess_enquiry_validity(email, str(row.get("Phone / WhatsApp", "")), details)
        try:
            project_amount = enquiry_store.parse_amount(row.get("Project Amount"))
            amount_received = enquiry_store.parse_amount(row.get("Amount Received")) or 0
            payment_date = parse_payment_date(row.get("Payment Date"))
            payment_status = safe_text(row.get("Payment Status"), 40) or "Pending"
            payment_due_date = parse_payment_date(row.get("Payment Due Date"))
            if payment_status not in {"Pending", "Partial", "Paid", "Refunded"}:
                raise ValueError("Payment Status is invalid")
            if amount_received and project_amount is None:
                raise ValueError("Project Amount is required before recording payment")
            if project_amount is not None and amount_received > project_amount:
                raise ValueError("Amount Received cannot exceed Project Amount")
            if amount_received and payment_date is None:
                raise ValueError("Payment Date is required when Amount Received is entered")
            if payment_status == "Paid" and project_amount is not None and amount_received != project_amount:
                raise ValueError("Paid status requires the full Project Amount")
            quote_defaults = quotation_values(enquiry_id, project_amount)
            quotation_amount = enquiry_store.parse_amount(row.get("Quotation Amount"))
            if quotation_amount is None:
                quotation_amount = project_amount
            delivery_days_text = str(row.get("Delivery Days") or quote_defaults["delivery_days"]).strip()
            delivery_days = int(float(delivery_days_text))
            if not 1 <= delivery_days <= 3650:
                raise ValueError("Delivery Days must be between 1 and 3650")
            valid_until = parse_payment_date(row.get("Valid Until")) or quote_defaults["quotation_valid_until"]
            quotation_status = safe_text(row.get("Quotation Status"), 40) or "Pending"
            if quotation_status not in {"Pending", "Sent", "Failed", "Accepted", "Rejected"}:
                raise ValueError("Quotation Status is invalid")
            thank_you_requested = str(row.get("Send Thank You") or "").strip().upper() in {
                "TRUE", "SEND", "YES", "1",
            }
            quotation_requested = str(row.get("Send Quotation") or "").strip().upper() in {
                "TRUE", "SEND", "YES", "1",
            }
            receipt_requested = str(row.get("Send Payment Receipt") or "").strip().upper() in {
                "TRUE", "SEND", "YES", "1",
            }
            reminder_requested = str(row.get("Send Payment Reminder") or "").strip().upper() in {
                "TRUE", "SEND", "YES", "1",
            }
            previous_received = enquiry_store.parse_amount(
                existing_record.get("amount_received") if existing_record else None
            ) or 0
            pending_amount = (
                max(project_amount - amount_received, 0)
                if project_amount is not None else None
            )
            payment_email_status = (
                safe_text(row.get("Payment Email Status"), 60)
                or safe_text(existing_record.get("payment_email_status") if existing_record else None, 60)
                or "Pending"
            )
            enquiry_store.upsert_from_sheet({
                "enquiry_id": enquiry_id,
                "received_at": datetime.now(timezone.utc),
                "client_name": name, "email": email,
                "phone": safe_text(row.get("Phone / WhatsApp"), 40),
                "project_subject": subject, "project_details": details,
                "project_amount": project_amount,
                "amount_received": amount_received,
                "payment_date": payment_date,
                "payment_status": payment_status,
                "lead_status": safe_text(row.get("Lead Status"), 40) or "New",
                "email_status": safe_text(row.get("Email Delivery"), 40) or "Manual",
                "source": safe_text(row.get("Source"), 80) or "Google Sheet",
                "validity": validity, "validation_notes": notes,
                "quotation_number": safe_text(row.get("Quotation Number"), 60) or quote_defaults["quotation_number"],
                "quotation_amount": quotation_amount,
                "quotation_currency": safe_text(row.get("Currency"), 10).upper() or "INR",
                "delivery_days": delivery_days,
                "quotation_valid_until": valid_until,
                "payment_terms": safe_text(row.get("Payment Terms"), 500) or quote_defaults["payment_terms"],
                "quotation_notes": safe_text(row.get("Quotation Notes"), 1000) or quote_defaults["quotation_notes"],
                "quotation_status": quotation_status,
                "quotation_sent_at": None,
                "payment_due_date": payment_due_date,
                "payment_email_status": payment_email_status,
                "payment_receipt_sent_at": existing_record.get("payment_receipt_sent_at") if existing_record else None,
                "payment_reminder_sent_at": existing_record.get("payment_reminder_sent_at") if existing_record else None,
            }, index)
            due_date_changed = bool(
                existing_record
                and existing_record.get("payment_due_date") != payment_due_date
            )
            if due_date_changed:
                enquiry_store.reset_payment_reminder(enquiry_id)
            saved_record = enquiry_store.get_enquiry(enquiry_id)
            if thank_you_requested and saved_record:
                thank_you_sent = send_enquiry_email(
                    enquiry_id=enquiry_id,
                    name=name,
                    email=email,
                    phone=safe_text(row.get("Phone / WhatsApp"), 40),
                    subject=subject,
                    message=details,
                    notify_owner=False,
                )
                delivery_status = "Sent" if thank_you_sent else "Failed"
                enquiry_store.update_delivery(enquiry_id, delivery_status)
                sheet.update_cell(index, SYNC_COLUMN["Email Delivery"], delivery_status)
                sheet.batch_clear([f"AB{index}"])
            if quotation_requested:
                quotation_sent = bool(saved_record) and send_quotation_email(saved_record)
                quotation_status = "Sent" if quotation_sent else "Failed"
                enquiry_store.update_quotation_delivery(enquiry_id, quotation_status)
                sheet.update_cell(index, SYNC_COLUMN["Quotation Status"], quotation_status)
                sheet.update_cell(
                    index, SYNC_COLUMN["Quotation Sent At"],
                    format_india_datetime() if quotation_sent else "",
                )
                sheet.batch_clear([f"AC{index}"])
            saved_record = enquiry_store.get_enquiry(enquiry_id)
            auto_receipt = amount_received > previous_received
            if saved_record and (receipt_requested or auto_receipt):
                receipt_sent = send_payment_email(
                    saved_record,
                    kind="receipt",
                    received_amount=(amount_received - previous_received) if auto_receipt else amount_received,
                )
                if receipt_sent:
                    enquiry_store.mark_payment_receipt_sent(enquiry_id)
                    payment_email_status = "Receipt Sent"
                    sheet.update_cell(index, SYNC_COLUMN["Payment Receipt Sent At"], format_india_datetime())
                else:
                    payment_email_status = "Receipt Failed"
                sheet.batch_clear([f"AI{index}"])
            overdue = bool(
                payment_due_date
                and payment_due_date < datetime.now(INDIA_TIMEZONE).date()
                and pending_amount is not None
                and pending_amount > 0
            )
            reminder_already_sent = bool(
                existing_record
                and existing_record.get("payment_reminder_sent_at")
                and not due_date_changed
            )
            if saved_record and (reminder_requested or (overdue and not reminder_already_sent)):
                reminder_sent = send_payment_email(saved_record, kind="reminder")
                if reminder_sent:
                    enquiry_store.mark_payment_reminder_sent(enquiry_id)
                    payment_email_status = "Reminder Sent"
                    sheet.update_cell(index, SYNC_COLUMN["Payment Reminder Sent At"], format_india_datetime())
                else:
                    payment_email_status = "Reminder Failed"
                sheet.batch_clear([f"AJ{index}"])
            sheet.update_cell(
                index, SYNC_COLUMN["Pending Amount"],
                float(pending_amount) if pending_amount is not None else "",
            )
            sheet.update_cell(index, SYNC_COLUMN["Payment Email Status"], payment_email_status)
            sheet.update_cell(index, SYNC_COLUMN["Updated At"], format_india_datetime())
            sheet.update_cell(index, SYNC_COLUMN["Sync Status"], "Synced")
            result["created_or_updated"] += 1
        except Exception as exc:
            app.logger.exception("Sheet row %s failed to sync", index)
            sheet.update_cell(index, SYNC_COLUMN["Sync Status"], f"Error: {type(exc).__name__}"[:100])
            result["skipped"] += 1
    return result


def send_enquiry_email(
    *,
    enquiry_id: str,
    name: str,
    email: str,
    phone: str,
    subject: str,
    message: str,
    notify_owner: bool = True,
) -> bool:
    if not all([MAIL_USERNAME, MAIL_PASSWORD, MAIL_RECEIVER]):
        app.logger.warning("Gmail settings are missing; enquiry saved to database only.")
        return False

    received_at = format_india_datetime()
    h_enquiry_id = html_lib.escape(enquiry_id)
    h_name = html_lib.escape(name)
    h_email = html_lib.escape(email)
    h_phone = html_lib.escape(phone or "Not provided")
    h_subject = html_lib.escape(subject)
    h_message = html_lib.escape(message)
    owner_mail = EmailMessage()
    owner_mail["Subject"] = f"[{enquiry_id}] New Portfolio Enquiry - {subject}"
    owner_mail["From"] = f"Shubham Chauhan Portfolio <{MAIL_USERNAME}>"
    owner_mail["To"] = MAIL_RECEIVER
    owner_mail["Reply-To"] = email
    owner_mail.set_content(
        f"""A new enquiry was submitted from the portfolio website.

CLIENT DETAILS
Enquiry ID: {enquiry_id}
Name: {name}
Email: {email}
Phone: {phone or "Not provided"}
Subject: {subject}

MESSAGE
{message}

Received: {received_at}
"""
    )
    owner_mail.add_alternative(
        f"""<!doctype html>
<html><body style="margin:0;background:#f1f5f9;font-family:Arial,sans-serif;color:#172554">
<div style="max-width:640px;margin:28px auto;background:#fff;border-radius:18px;overflow:hidden;border:1px solid #dbeafe">
  <div style="padding:24px 28px;background:linear-gradient(135deg,#172554,#2563eb);color:#fff">
    <img src="cid:sc-owner-logo" alt="Shubham Chauhan" width="72" height="72" style="display:block;border-radius:14px;margin-bottom:15px;background:#fff">
    <div style="font-size:12px;letter-spacing:2px;font-weight:700">SHUBHAM CHAUHAN · SOFTWARE ENGINEER</div>
    <h1 style="margin:10px 0 0;font-size:24px">New portfolio enquiry</h1>
  </div>
  <div style="padding:28px">
    <p style="margin-top:0;color:#64748b">A new project enquiry has arrived. Reply directly to this email to contact the client.</p>
    <div style="margin:0 0 18px;padding:14px 16px;border-radius:12px;background:#172554;color:#fff"><span style="font-size:11px;letter-spacing:1px;color:#bfdbfe">ENQUIRY ID</span><br><strong style="font-size:20px">{h_enquiry_id}</strong></div>
    <table role="presentation" style="width:100%;border-collapse:collapse;font-size:14px">
      <tr><td style="padding:9px 0;color:#64748b;width:145px">Client</td><td style="padding:9px 0;font-weight:700">{h_name}</td></tr>
      <tr><td style="padding:9px 0;color:#64748b">Email</td><td style="padding:9px 0">{h_email}</td></tr>
      <tr><td style="padding:9px 0;color:#64748b">Phone / WhatsApp</td><td style="padding:9px 0">{h_phone}</td></tr>
      <tr><td style="padding:9px 0;color:#64748b">Project</td><td style="padding:9px 0;font-weight:700">{h_subject}</td></tr>
    </table>
    <div style="margin-top:18px;padding:18px;border-radius:12px;background:#eff6ff;border-left:4px solid #2563eb;white-space:pre-wrap">{h_message}</div>
    <p style="margin:20px 0 0;color:#94a3b8;font-size:12px">Received: {received_at}</p>
  </div>
</div></body></html>""",
        subtype="html",
    )
    if LOGO_PATH.exists():
        owner_mail.get_payload()[-1].add_related(
            LOGO_PATH.read_bytes(),
            maintype="image",
            subtype="png",
            cid="<sc-owner-logo>",
            filename="shubham-chauhan-logo.png",
        )

    reply_mail = EmailMessage()
    reply_mail["Subject"] = f"[{enquiry_id}] Your project enquiry is confirmed"
    reply_mail["From"] = f"Shubham Chauhan <{MAIL_USERNAME}>"
    reply_mail["To"] = email
    reply_mail["Reply-To"] = MAIL_RECEIVER
    reply_mail.set_content(
        f"""Hello {name},

Thank you for choosing to discuss your project with Shubham Chauhan.

Your enquiry regarding "{subject}" has been received successfully. Your requirements will now be reviewed for scope, technical approach, delivery plan, and commercial estimate. A separate quotation email will follow with the available project details.

Your Enquiry ID: {enquiry_id}
Please keep this ID and include it in every future requirement, change request or issue related to this project.

Your message:
{message}

Thank you,
Shubham Chauhan
Full Stack Developer
Web Apps | Mobile Apps | APIs | Cloud Solutions
"""
    )
    reply_mail.add_alternative(
        f"""<!doctype html>
<html><body style="margin:0;background:#f1f5f9;font-family:Arial,sans-serif;color:#0f172a">
<div style="max-width:640px;margin:28px auto;background:#fff;border-radius:20px;overflow:hidden;border:1px solid #dbeafe;box-shadow:0 16px 40px rgba(15,23,42,.08)">
  <div style="padding:26px 30px;background:linear-gradient(135deg,#172554,#2563eb);color:#fff">
    <img src="cid:sc-client-logo" alt="Shubham Chauhan" width="82" height="82" style="display:block;border-radius:16px;margin-bottom:16px;background:#fff">
    <div style="font-size:12px;letter-spacing:2px;font-weight:700">SC · SHUBHAM CHAUHAN</div>
    <h1 style="margin:11px 0 5px;font-size:25px">Thank you for reaching out.</h1>
    <p style="margin:0;color:#dbeafe;font-size:14px">Your project enquiry has been received successfully.</p>
  </div>
  <div style="padding:30px">
    <p style="font-size:16px">Hello <strong>{h_name}</strong>,</p>
    <p style="line-height:1.75;color:#475569">Thank you for sharing your requirements for <strong>{h_subject}</strong>. Your enquiry has been securely recorded and will be reviewed for scope, technical approach, delivery plan, and commercial estimate.</p>
    <div style="margin:20px 0;padding:18px;border-radius:14px;background:#f8fafc;border:1px solid #e2e8f0">
      <div style="margin-bottom:10px;color:#172554;font-size:12px;font-weight:800;letter-spacing:1px">WHAT HAPPENS NEXT</div>
      <div style="line-height:1.8;color:#475569">1. Your submitted requirements are reviewed.<br>2. A separate quotation email is prepared and sent.<br>3. You can reply to either email for clarification, changes, or approval.</div>
    </div>
    <div style="margin:20px 0;padding:18px;border-radius:14px;background:#172554;color:#fff;text-align:center">
      <div style="font-size:11px;letter-spacing:1.5px;color:#bfdbfe;font-weight:700">YOUR ENQUIRY ID</div>
      <div style="margin-top:7px;font-size:24px;font-weight:800;letter-spacing:.5px">{h_enquiry_id}</div>
      <div style="margin-top:9px;color:#dbeafe;font-size:13px;line-height:1.5">Keep this ID safe. Mention it in every future requirement, change request or issue so your project can be found immediately.</div>
    </div>
    <div style="margin:22px 0;padding:18px;border-radius:14px;background:#eff6ff;border:1px solid #dbeafe">
      <div style="margin-bottom:8px;color:#2563eb;font-size:12px;font-weight:700;letter-spacing:1px">YOUR ENQUIRY</div>
      <div style="white-space:pre-wrap;line-height:1.65;color:#334155">{h_message}</div>
    </div>
    <p style="line-height:1.65">Thank you,<br><strong style="color:#172554">Shubham Chauhan</strong><br><span style="color:#64748b">Full Stack Developer · Web Apps · Mobile Apps · APIs · Cloud Solutions</span></p>
  </div>
  <div style="padding:16px 30px;background:#f8fafc;color:#94a3b8;font-size:11px;text-align:center">This is an automatic confirmation from the Shubham Chauhan portfolio.</div>
</div></body></html>""",
        subtype="html",
    )
    if LOGO_PATH.exists():
        reply_mail.get_payload()[-1].add_related(
            LOGO_PATH.read_bytes(),
            maintype="image",
            subtype="png",
            cid="<sc-client-logo>",
            filename="shubham-chauhan-logo.png",
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as smtp:
        smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
        if notify_owner:
            smtp.send_message(owner_mail)
        smtp.send_message(reply_mail)

    return True


def send_quotation_email(record: dict[str, Any]) -> bool:
    """Send the client a separate quotation after their confirmation email."""
    if not all([MAIL_USERNAME, MAIL_PASSWORD, MAIL_RECEIVER]):
        app.logger.warning("Gmail settings are missing; quotation was not sent.")
        return False

    enquiry_id = safe_text(record.get("enquiry_id"), 40)
    quote = quotation_values(enquiry_id, record.get("project_amount"))
    for key in quote:
        if record.get(key) not in (None, ""):
            quote[key] = record[key]

    name = safe_text(record.get("client_name"), 100)
    email = safe_text(record.get("email"), 180)
    subject = safe_text(record.get("project_subject"), 160)
    details = safe_text(record.get("project_details"), 5000)
    amount = quote["quotation_amount"]
    amount_text = (
        f"{quote['quotation_currency']} {float(amount):,.2f}"
        if amount is not None else "To be finalized after requirement discussion"
    )
    valid_until = quote["quotation_valid_until"]
    if isinstance(valid_until, str):
        valid_until_text = valid_until
    else:
        valid_until_text = valid_until.strftime("%d %b %Y")

    mail = EmailMessage()
    mail["Subject"] = f"[{quote['quotation_number']}] Project quotation - {subject}"
    mail["From"] = f"Shubham Chauhan <{MAIL_USERNAME}>"
    mail["To"] = email
    mail["Reply-To"] = MAIL_RECEIVER
    mail.set_content(
        f"""Hello {name},

Thank you for the opportunity to understand your project. Based on the requirements currently available, I have prepared the following preliminary quotation for your review.

Quotation: {quote['quotation_number']}
Enquiry: {enquiry_id}
Project: {subject}
Amount: {amount_text}
Delivery estimate: {quote['delivery_days']} days
Valid until: {valid_until_text}
Payment terms: {quote['payment_terms']}

Scope:
{details}

Notes:
{quote['quotation_notes']}

This quotation provides a clear starting point. Final scope, milestones, integrations, and delivery dates will be confirmed before development begins.

Reply to this email with "Approved" to proceed, or share any changes you would like included.

Thank you,
Shubham Chauhan
Full Stack Developer
"""
    )
    escaped = {key: html_lib.escape(str(value)) for key, value in {
        "name": name, "number": quote["quotation_number"], "enquiry": enquiry_id,
        "subject": subject, "amount": amount_text, "days": quote["delivery_days"],
        "valid": valid_until_text, "terms": quote["payment_terms"],
        "details": details, "notes": quote["quotation_notes"],
    }.items()}
    mail.add_alternative(
        f"""<!doctype html><html><body style="margin:0;background:#f1f5f9;font-family:Arial,sans-serif;color:#0f172a">
<div style="max-width:680px;margin:28px auto;background:#fff;border:1px solid #dbeafe;border-radius:20px;overflow:hidden">
  <div style="padding:26px 30px;background:linear-gradient(135deg,#172554,#2563eb);color:#fff">
    <div style="font-size:12px;letter-spacing:2px;font-weight:700">PROJECT QUOTATION</div>
    <h1 style="margin:10px 0 4px;font-size:25px">{escaped['number']}</h1>
    <p style="margin:0;color:#dbeafe">Prepared for {escaped['name']}</p>
  </div>
  <div style="padding:30px">
    <p style="margin-top:0;line-height:1.75;color:#475569">Thank you for the opportunity to understand your project. Based on the requirements currently available, I have prepared this preliminary quotation as a clear starting point for our discussion.</p>
    <table role="presentation" style="width:100%;border-collapse:collapse;font-size:14px">
      <tr><td style="padding:9px;color:#64748b">Enquiry ID</td><td style="padding:9px;font-weight:700">{escaped['enquiry']}</td></tr>
      <tr><td style="padding:9px;color:#64748b">Project</td><td style="padding:9px;font-weight:700">{escaped['subject']}</td></tr>
      <tr><td style="padding:9px;color:#64748b">Quotation amount</td><td style="padding:9px;font-size:19px;font-weight:800;color:#2563eb">{escaped['amount']}</td></tr>
      <tr><td style="padding:9px;color:#64748b">Delivery estimate</td><td style="padding:9px">{escaped['days']} days</td></tr>
      <tr><td style="padding:9px;color:#64748b">Valid until</td><td style="padding:9px">{escaped['valid']}</td></tr>
      <tr><td style="padding:9px;color:#64748b">Payment terms</td><td style="padding:9px">{escaped['terms']}</td></tr>
    </table>
    <div style="margin-top:20px;padding:18px;border-radius:14px;background:#eff6ff"><strong>Project scope</strong><div style="margin-top:8px;white-space:pre-wrap;line-height:1.65">{escaped['details']}</div></div>
    <div style="margin-top:18px;padding:16px 18px;border-radius:12px;background:#fffbeb;border:1px solid #fde68a;color:#78350f;line-height:1.65"><strong>Important note</strong><br>{escaped['notes']}</div>
    <p style="margin-top:24px;line-height:1.7">Final scope, milestones, integrations, and delivery dates will be confirmed before development begins. Reply with <strong>Approved</strong> to proceed, or share the changes you would like included.</p>
    <p style="line-height:1.65">Thank you,<br><strong style="color:#172554">Shubham Chauhan</strong><br><span style="color:#64748b">Full Stack Developer · Web Apps · Mobile Apps · APIs · Cloud Solutions</span></p>
  </div>
</div></body></html>""",
        subtype="html",
    )
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as smtp:
        smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
        smtp.send_message(mail)
    return True


def send_payment_email(
    record: dict[str, Any], *, kind: str, received_amount: Any = None
) -> bool:
    """Send a professional payment receipt or overdue-balance reminder."""
    if kind not in {"receipt", "reminder"}:
        raise ValueError("Unsupported payment email type")
    if not all([MAIL_USERNAME, MAIL_PASSWORD, MAIL_RECEIVER]):
        app.logger.warning("Gmail settings are missing; payment email was not sent.")
        return False

    enquiry_id = safe_text(record.get("enquiry_id"), 40)
    name = safe_text(record.get("client_name"), 100)
    email = safe_text(record.get("email"), 180)
    project = safe_text(record.get("project_subject"), 160)
    total = enquiry_store.parse_amount(record.get("project_amount"))
    paid = enquiry_store.parse_amount(record.get("amount_received")) or 0
    pending = max((total or 0) - paid, 0) if total is not None else None
    received = enquiry_store.parse_amount(received_amount)
    currency = safe_text(record.get("quotation_currency"), 10) or "INR"
    due_date = record.get("payment_due_date")
    due_text = (
        due_date.strftime("%d %b %Y") if isinstance(due_date, date)
        else safe_text(due_date, 40) or "Not specified"
    )

    def money(value: Any) -> str:
        return f"{currency} {float(value):,.2f}" if value is not None else "To be confirmed"

    is_receipt = kind == "receipt"
    title = "Payment received successfully" if is_receipt else "Friendly payment reminder"
    subject = (
        f"[{enquiry_id}] Payment received - {project}"
        if is_receipt else f"[{enquiry_id}] Pending payment reminder - {project}"
    )
    intro = (
        "Thank you. Your payment has been received and recorded successfully. "
        "Please keep this email as confirmation for your project records."
        if is_receipt else
        "This is a courteous reminder that a payment balance remains pending for your project. "
        "If payment has already been completed, please reply with the transaction reference so the account can be updated."
    )
    action = (
        "No action is required for the amount already received. The remaining balance will follow the agreed project milestones."
        if is_receipt else
        "Please arrange the pending payment at your earliest convenience, or reply if you need any clarification regarding the amount or due date."
    )
    received_line = money(received if received is not None else paid)
    mail = EmailMessage()
    mail["Subject"] = subject
    mail["From"] = f"Shubham Chauhan <{MAIL_USERNAME}>"
    mail["To"] = email
    mail["Reply-To"] = MAIL_RECEIVER
    mail.set_content(
        f"""Hello {name},

{intro}

Project: {project}
Enquiry ID: {enquiry_id}
Payment received: {received_line if is_receipt else money(paid)}
Total project amount: {money(total)}
Pending balance: {money(pending)}
Payment due date: {due_text}

{action}

Thank you,
Shubham Chauhan
Full Stack Developer
Web Apps | Mobile Apps | APIs | Cloud Solutions
"""
    )
    escaped = {key: html_lib.escape(str(value)) for key, value in {
        "name": name, "title": title, "intro": intro, "project": project,
        "enquiry": enquiry_id, "received": received_line if is_receipt else money(paid),
        "total": money(total), "pending": money(pending), "due": due_text,
        "action": action,
    }.items()}
    mail.add_alternative(
        f"""<!doctype html><html><body style="margin:0;background:#f4f7fb;font-family:Arial,sans-serif;color:#0f172a">
<div style="max-width:680px;margin:28px auto;background:#fff;border:1px solid #dbeafe;border-radius:20px;overflow:hidden">
  <div style="padding:28px 32px;background:#123b8f;color:#fff">
    <div style="font-size:12px;letter-spacing:2px;font-weight:800">PAYMENT UPDATE</div>
    <h1 style="margin:10px 0 4px;font-size:25px">{escaped['title']}</h1>
    <p style="margin:0;color:#dbeafe">{escaped['project']} · {escaped['enquiry']}</p>
  </div>
  <div style="padding:32px">
    <p style="font-size:16px">Hello <strong>{escaped['name']}</strong>,</p>
    <p style="line-height:1.75;color:#475569">{escaped['intro']}</p>
    <table role="presentation" style="width:100%;border-collapse:collapse;margin:22px 0;font-size:14px">
      <tr style="background:#eff6ff"><td style="padding:12px;color:#64748b">Payment received</td><td style="padding:12px;text-align:right;font-weight:800;color:#1d4ed8">{escaped['received']}</td></tr>
      <tr><td style="padding:12px;color:#64748b">Total project amount</td><td style="padding:12px;text-align:right;font-weight:700">{escaped['total']}</td></tr>
      <tr style="background:#eff6ff"><td style="padding:12px;color:#64748b">Pending balance</td><td style="padding:12px;text-align:right;font-size:18px;font-weight:800;color:#123b8f">{escaped['pending']}</td></tr>
      <tr><td style="padding:12px;color:#64748b">Payment due date</td><td style="padding:12px;text-align:right;font-weight:700">{escaped['due']}</td></tr>
    </table>
    <div style="padding:17px 19px;border-radius:12px;background:#f8fafc;border-left:4px solid #2563eb;line-height:1.7;color:#334155">{escaped['action']}</div>
    <p style="margin-top:26px;line-height:1.65">Thank you,<br><strong style="color:#123b8f">Shubham Chauhan</strong><br><span style="color:#64748b">Full Stack Developer · Web Apps · Mobile Apps · APIs · Cloud Solutions</span></p>
  </div>
</div></body></html>""",
        subtype="html",
    )
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as smtp:
        smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
        smtp.send_message(mail)
    return True


def local_assistant_reply(message: str) -> str:
    """Useful fallback when no OpenAI key is configured or the API is unavailable."""
    lowered = message.lower()

    if any(term in lowered for term in ("e-com", "ecom", "ecommerce", "e-commerce", "shop")):
        return (
            "Shubham can build a single-vendor or multi-vendor e-commerce platform "
            "with Flutter apps, a responsive website, product management, cart, "
            "payments, order tracking, inventory and an admin dashboard. "
            "Do you need a single store or a marketplace with multiple sellers?"
        )

    if any(term in lowered for term in ("dating", "matchmaking", "matrimonial")):
        return (
            "A dating platform can include profile verification, preferences, swipe "
            "matching, chat, subscriptions, package limits, payments, report/block "
            "controls and an admin moderation panel. Which packages and platforms "
            "do you want: Android, iOS, website, or all three?"
        )

    if any(term in lowered for term in ("school", "erp", "inventory", "crm")):
        return (
            "Shubham develops role-based ERP systems with dashboards, permissions, "
            "reports, notifications, mobile apps and Python APIs. Tell me the user "
            "roles and the most important modules for your system."
        )

    if any(term in lowered for term in ("flutter", "android", "ios", "mobile app")):
        return (
            "Shubham builds Flutter apps for Android and iOS with authentication, "
            "Firebase, REST APIs, payments, maps, notifications and admin panels. "
            "Please describe your main user journey."
        )

    if any(term in lowered for term in ("website", "portfolio", "web app")):
        return (
            "Shubham can build a responsive website or web application with Python, "
            "Flask or FastAPI, dynamic content, database, admin panel, SEO and enquiry "
            "automation. What type of business is the website for?"
        )

    return (
        "I can help you plan mobile apps, websites, e-commerce platforms, ERP systems, "
        "Python APIs and admin dashboards. Share your project idea, required platforms "
        "and key features, and I will structure the solution."
    )


def sanitise_history(history: Any) -> list[dict[str, str]]:
    clean: list[dict[str, str]] = []
    if not isinstance(history, list):
        return clean

    for item in history[-14:]:
        if not isinstance(item, dict):
            continue
        role = item.get("role")
        content = safe_text(item.get("content"), 4000)
        if role in {"user", "assistant"} and content:
            clean.append({"role": role, "content": content})
    return clean


ALLOWED_CHAT_FILES = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "application/pdf": ".pdf",
}


def read_chat_attachments() -> tuple[list[dict[str, Any]], str | None]:
    attachments: list[dict[str, Any]] = []
    total_size = 0
    for uploaded in request.files.getlist("attachments")[:3]:
        content_type = (uploaded.mimetype or "").lower()
        if content_type not in ALLOWED_CHAT_FILES:
            return [], "Only JPG, PNG, WebP and PDF files are supported."
        data = uploaded.read(5 * 1024 * 1024 + 1)
        if len(data) > 5 * 1024 * 1024:
            return [], "Each attachment must be 5 MB or smaller."
        total_size += len(data)
        if total_size > 10 * 1024 * 1024:
            return [], "Attachments must be 10 MB or smaller in total."
        attachments.append(
            {
                "name": safe_text(uploaded.filename, 180) or f"attachment{ALLOWED_CHAT_FILES[content_type]}",
                "content_type": content_type,
                "data": data,
            }
        )
    return attachments, None


def send_chat_copy(message: str, attachments: list[dict[str, Any]]) -> None:
    """Forward each visitor chat message and its files to the portfolio owner."""
    if not all([MAIL_USERNAME, MAIL_PASSWORD, MAIL_RECEIVER]):
        return
    mail = EmailMessage()
    mail["Subject"] = "Portfolio AI Chat - New visitor message"
    mail["From"] = f"Shubham Chauhan Portfolio <{MAIL_USERNAME}>"
    mail["To"] = MAIL_RECEIVER
    file_names = ", ".join(item["name"] for item in attachments) or "None"
    mail.set_content(
        f"Visitor message:\n\n{message or 'Attachment shared without text'}\n\n"
        f"Attachments: {file_names}\nReceived: {format_india_datetime()}"
    )
    mail.add_alternative(
        f"""<!doctype html><html><body style="margin:0;background:#f1f5f9;font-family:Arial,sans-serif;color:#0f172a">
<div style="max-width:620px;margin:28px auto;background:#fff;border:1px solid #dbeafe;border-radius:18px;overflow:hidden">
<div style="padding:24px;background:linear-gradient(135deg,#172554,#2563eb);color:#fff">
<img src="cid:sc-chat-logo" width="72" height="72" alt="Shubham Chauhan" style="display:block;border-radius:14px;background:#fff;margin-bottom:14px">
<strong style="font-size:20px">New AI chat message</strong></div>
<div style="padding:26px"><p style="white-space:pre-wrap;line-height:1.7">{html_lib.escape(message or 'Attachment shared without text')}</p>
<p style="padding:14px;background:#eff6ff;border-radius:12px"><strong>Attachments:</strong> {html_lib.escape(file_names)}</p>
<p style="color:#64748b;font-size:12px">Received: {format_india_datetime()}</p></div></div></body></html>""",
        subtype="html",
    )
    if LOGO_PATH.exists():
        mail.get_payload()[-1].add_related(
            LOGO_PATH.read_bytes(), maintype="image", subtype="png",
            cid="<sc-chat-logo>", filename="shubham-chauhan-logo.png",
        )
    for item in attachments:
        maintype, subtype = item["content_type"].split("/", 1)
        mail.add_attachment(
            item["data"],
            maintype=maintype,
            subtype=subtype,
            filename=item["name"],
        )
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as smtp:
        smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
        smtp.send_message(mail)


def forward_chat_copy_safely(message: str, attachments: list[dict[str, Any]]) -> None:
    try:
        send_chat_copy(message, attachments)
    except Exception:
        app.logger.exception("Unable to forward the visitor chat message.")


@app.context_processor
def global_template_values() -> dict[str, Any]:
    whatsapp_message = (
        "Hello Shubham, I visited your portfolio and want to discuss a project."
    )
    whatsapp_url = (
        f"https://wa.me/{WHATSAPP_NUMBER}"
        f"?text={whatsapp_message.replace(' ', '%20')}"
    )
    return {
        "current_year": datetime.now().year,
        "whatsapp_url": whatsapp_url,
        "ai_enabled": bool(openai_client),
    }


@app.get("/")
def home() -> str:
    projects = [
        {
            "icon": "fa-layer-group",
            "category": "DIGITAL PLATFORM",
            "title": "WASEF",
            "image": "images/projects/wasef-logo.jpeg",
            "description": "A structured digital product focused on a clear, reliable and responsive user experience.",
            "tags": ["Product", "Mobile", "API"],
        },
        {
            "icon": "fa-chart-line",
            "category": "BUSINESS OPERATIONS",
            "title": "Biz-Control",
            "image": "images/projects/biz-control-logo.jpeg",
            "description": "A business-control solution designed to organise day-to-day operations and management workflows.",
            "tags": ["Business", "Dashboard", "Workflow"],
            "store_live": False,
        },
        {
            "icon": "fa-motorcycle",
            "category": "COLLECTION PLATFORM",
            "title": "CollectX Agency",
            "image": "images/projects/collectx-agency-icon.png",
            "description": "The agency-facing CollectX application for coordinated collection and operational workflows.",
            "tags": ["Agency App", "Collections", "Operations"],
            "store_url": "https://play.google.com/store/apps/details?id=com.arramton.collectx&hl=en_IN",
            "store_live": True,
        },
        {
            "icon": "fa-motorcycle",
            "category": "COLLECTION PLATFORM",
            "title": "CollectX Rider",
            "image": "images/projects/collectx-rider-icon.png",
            "description": "The rider application supporting assigned collection tasks and field operations.",
            "tags": ["Rider App", "Field Work", "Operations"],
            "store_url": "https://play.google.com/store/apps/details?id=com.arramton.collectxrider&hl=en_IN",
            "store_live": True,
        },
        {
            "icon": "fa-shirt",
            "category": "CLOTHING STORAGE",
            "title": "Closet",
            "description": "A customer application designed to organise wardrobe storage and clothing-management requests.",
            "tags": ["User App", "Wardrobe", "Storage"],
            "store_live": False,
        },
        {
            "icon": "fa-truck-fast",
            "category": "CLOTHING STORAGE",
            "title": "Closet Rider",
            "description": "A field-team application for clothing pickups, task updates and operational coordination.",
            "tags": ["Rider App", "Pickups", "Operations"],
            "store_live": False,
        },
        {
            "icon": "fa-cake-candles",
            "category": "FOOD & COMMERCE",
            "title": "Cakingom - Online Cake Delivery",
            "image": "images/projects/cakingom-icon.png",
            "description": "Cake, flowers, personalised gifts and celebration-decoration ordering with doorstep delivery.",
            "tags": ["Shopping", "Orders", "Delivery"],
            "store_url": "https://play.google.com/store/apps/details?id=com.arramton.cakingom",
            "store_live": True,
        },
        {
            "icon": "fa-store",
            "category": "BUSINESS",
            "title": "Keyenzy",
            "image": "images/projects/keyenzy-icon.png",
            "description": "A mobile business application that helps registered shop owners manage their online presence and operations.",
            "tags": ["Shop Owners", "Business", "Management"],
            "store_url": "https://play.google.com/store/apps/details?id=com.arramton.keyenzy",
            "store_live": True,
        },
        {
            "icon": "fa-car-side",
            "category": "AUTO MARKETPLACE",
            "title": "Carenzy",
            "image": "images/projects/carenzy-icon.png",
            "description": "A used-car marketplace connecting buyers, sellers and dealers through listings, bidding and lead workflows.",
            "tags": ["Marketplace", "Bidding", "Automotive"],
            "store_url": "https://play.google.com/store/apps/details?id=com.car.carenzy",
            "store_live": True,
        },
        {
            "icon": "fa-key",
            "category": "BUSINESS ORDERING",
            "title": "KMD Keys",
            "image": "images/projects/kmd-keys-icon.png",
            "description": "A specialised ordering application created for keymakers to submit orders for faster delivery.",
            "tags": ["Orders", "Keymakers", "Delivery"],
            "store_url": "https://play.google.com/store/apps/details?id=com.app.kmdkeys",
            "store_live": True,
        },
        {
            "icon": "fa-print",
            "category": "PRINT SERVICES",
            "title": "ZoomPrint",
            "description": "A mobile product centred on convenient digital print-service requests and workflow coordination.",
            "tags": ["Printing", "Orders", "Services"],
            "store_live": False,
        },
        {
            "icon": "fa-house-circle-check",
            "category": "HOME SERVICES",
            "title": "The NinjaCare",
            "image": "images/projects/ninjacare-user-icon.png",
            "description": "A customer app for scheduling household cleaning, cooking and domestic-helper services.",
            "tags": ["User App", "Bookings", "Home Services"],
            "store_live": False,
        },
        {
            "icon": "fa-people-carry-box",
            "category": "HOME SERVICES",
            "title": "NinjaCare Partner",
            "image": "images/projects/ninjacare-partner-icon.png",
            "description": "The partner-side experience supporting assigned household-service work and delivery operations.",
            "tags": ["Partner App", "Services", "Operations"],
            "store_live": False,
        },
        {
            "icon": "fa-burger",
            "category": "FOOD DELIVERY",
            "title": "Food Go",
            "description": "A food-ordering product focused on a quick customer journey and organised fulfilment workflow.",
            "tags": ["Food", "Delivery", "Orders"],
        },
        {
            "icon": "fa-plane-departure",
            "category": "TRAVEL",
            "title": "Flight Booking",
            "description": "A responsive flight-search and booking experience with a clear, traveller-friendly journey.",
            "tags": ["Travel", "Booking", "API"],
        },
        {
            "icon": "fa-user-graduate",
            "category": "EDUCATION MANAGEMENT",
            "title": "Student Management",
            "description": "A structured student-management system for organising academic and administrative information.",
            "tags": ["Education", "Management", "Dashboard"],
        },
        {
            "icon": "fa-screwdriver-wrench",
            "category": "SERVICE PLATFORM",
            "title": "Metri Serv",
            "description": "A service-oriented digital product designed to connect requests with an efficient operational workflow.",
            "tags": ["Services", "Operations", "Platform"],
        },
    ]

    reviews = [
        {
            "company": "WASEF",
            "image": "images/projects/wasef-logo.jpeg",
            "rating": 5,
            "text": "Shubham understood our requirements and presented the WASEF project in a clean, professional way. He paid attention to our feedback, made the requested improvements carefully, and produced work that matched the direction we had discussed.",
        },
        {
            "company": "Biz-Control",
            "image": "images/projects/biz-control-logo.jpeg",
            "rating": 4,
            "text": "Our experience with Shubham was positive and straightforward. He listened to the needs of Biz-Control, organised the work clearly, and created a practical final result. The quality of the completed project was very good.",
        },
        {
            "company": "CollectX Agency & Rider",
            "image": "images/projects/collectx-agency-icon.png",
            "rating": 5,
            "text": "The CollectX work was handled with care from the initial discussion through the final updates. Shubham remained cooperative, understood the separate agency and rider requirements, and completed the project in a polished and well-organised manner.",
        },
        {
            "company": "Keyenzy",
            "image": "images/projects/keyenzy-icon.png",
            "rating": 4,
            "text": "We appreciated Shubham’s patient and supportive approach while working on Keyenzy. Our ideas were considered properly, changes were addressed without confusion, and the finished application looks clean, useful, and professionally prepared.",
        },
        {
            "company": "Cakingom",
            "image": "images/projects/cakingom-icon.png",
            "rating": 5,
            "text": "Working with Shubham on Cakingom felt smooth and comfortable. He communicated clearly, gave proper attention to the details we shared, and refined the project based on our suggestions. We are happy with the overall outcome.",
        },
        {
            "company": "The NinjaCare",
            "image": "images/projects/ninjacare-user-icon.png",
            "rating": 5,
            "text": "Good work was delivered for The NinjaCare project. Shubham took time to understand what was needed for the customer and partner sides, responded well to feedback, and ensured the final presentation was clear and consistent.",
        },
    ]

    return render_template(
        "index.html",
        projects=projects,
        reviews=reviews,
        submission_token=str(uuid.uuid4()),
    )


@app.post("/contact")
def contact() -> Response:
    if safe_text(request.form.get("company_website"), 200):
        flash("Thank you. Your enquiry was received successfully.", "success")
        return redirect(url_for("home", _anchor="contact"))

    name = safe_text(request.form.get("name"), 100)
    email = safe_text(request.form.get("email"), 180)
    phone = safe_text(request.form.get("phone"), 40)
    subject = safe_text(request.form.get("subject"), 160)
    message = safe_text(request.form.get("message"), 5000)
    submission_token = safe_text(request.form.get("submission_token"), 36)

    if not all([name, email, subject, message]):
        flash("Please complete all required fields.", "error")
        return redirect(url_for("home", _anchor="contact"))

    if not looks_like_email(email):
        flash("Please enter a valid email address.", "error")
        return redirect(url_for("home", _anchor="contact"))

    try:
        project_amount = enquiry_store.parse_amount(request.form.get("project_amount"))
        uuid.UUID(submission_token)
    except (ValueError, AttributeError):
        flash("Please refresh the page and enter a valid project amount.", "error")
        return redirect(url_for("home", _anchor="contact"))

    created_at = datetime.now(INDIA_TIMEZONE).isoformat(timespec="seconds")

    if enquiry_store.configured():
        validity, validation_notes = assess_enquiry_validity(email, phone, message)
        try:
            enquiry_id = f"ENQ-{uuid.uuid4().hex[:12].upper()}"
            record, created = enquiry_store.create_enquiry({
                "enquiry_id": enquiry_id,
                "submission_token": submission_token,
                "received_at": created_at,
                "client_name": name,
                "email": email,
                "phone": phone,
                "project_subject": subject,
                "project_details": message,
                "project_amount": project_amount,
                "validity": validity,
                "validation_notes": validation_notes,
                **quotation_values(enquiry_id, project_amount),
            })
        except Exception:
            app.logger.exception("Unable to save enquiry to Postgres")
            flash("We could not save your enquiry. Please try again or use WhatsApp.", "error")
            return redirect(url_for("home", _anchor="contact"))

        if not created:
            flash("This enquiry was already received; no duplicate was created.", "success")
            return redirect(url_for("home", _anchor="contact"))

        email_sent = False
        try:
            email_sent = send_enquiry_email(
                enquiry_id=record["enquiry_id"],
                name=name, email=email, phone=phone, subject=subject, message=message
            )
        except Exception:
            app.logger.exception("Unable to deliver enquiry email")
        enquiry_store.update_delivery(
            record["enquiry_id"], "Sent" if email_sent else "Saved only"
        )
        quotation_sent = False
        if email_sent:
            try:
                quotation_sent = send_quotation_email(record)
            except Exception:
                app.logger.exception("Unable to deliver automatic quotation")
        enquiry_store.update_quotation_delivery(
            record["enquiry_id"], "Sent" if quotation_sent else "Pending"
        )
        record = enquiry_store.get_enquiry(record["enquiry_id"]) or record
        try:
            upsert_enquiry_to_google_sheet(record)
        except Exception:
            app.logger.exception("Unable to sync enquiry to Google Sheet")

        flash(
            "Thank you. Your confirmation and quotation were sent successfully."
            if email_sent and quotation_sent
            else "Thank you. Your enquiry was sent; the quotation is pending."
            if email_sent
            else "Your enquiry was saved; email delivery will be retried.",
            "success" if email_sent and quotation_sent else "warning",
        )
        return redirect(url_for("home", _anchor="contact"))

    initialise_database()
    with database_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO contact_messages
            (name, email, phone, subject, message, email_status, created_at)
            VALUES (?, ?, ?, ?, ?, 'pending', ?)
            """,
            (name, email, phone, subject, message, created_at),
        )
        message_id = cursor.lastrowid
        connection.commit()

    email_sent = False
    try:
        email_sent = send_enquiry_email(
            enquiry_id=f"ENQ-{int(message_id):05d}",
            name=name,
            email=email,
            phone=phone,
            subject=subject,
            message=message,
        )
    except smtplib.SMTPAuthenticationError:
        app.logger.exception("Gmail authentication failed.")
    except Exception:
        app.logger.exception("Unable to deliver enquiry email.")

    quotation_sent = False
    if email_sent:
        try:
            quotation_sent = send_quotation_email({
                "enquiry_id": f"ENQ-{int(message_id):05d}",
                "client_name": name,
                "email": email,
                "project_subject": subject,
                "project_details": message,
                "project_amount": project_amount,
            })
        except Exception:
            app.logger.exception("Unable to deliver automatic quotation.")

    with database_connection() as connection:
        connection.execute(
            "UPDATE contact_messages SET email_status = ? WHERE id = ?",
            ("sent" if email_sent else "saved_only", message_id),
        )
        connection.commit()

    if not os.getenv("VERCEL"):
        try:
            append_enquiry_to_excel(
                enquiry_id=int(message_id),
                created_at=created_at,
                name=name,
                email=email,
                phone=phone,
                subject=subject,
                message=message,
                email_status="sent" if email_sent else "saved_only",
            )
        except Exception:
            app.logger.exception("Unable to update the Excel enquiry register.")

    try:
        append_enquiry_to_google_sheet(
            enquiry_id=int(message_id),
            created_at=created_at,
            name=name,
            email=email,
            phone=phone,
            subject=subject,
            message=message,
            email_status="sent" if email_sent else "saved_only",
        )
    except Exception:
        app.logger.exception("Unable to update the Google Sheet enquiry register.")

    if email_sent and quotation_sent:
        flash("Thank you. Your confirmation and quotation were sent successfully.", "success")
    elif email_sent:
        flash("Thank you. Your confirmation was sent; the quotation is pending.", "warning")
    else:
        flash(
            "Your enquiry was saved. Email delivery is not configured yet, "
            "so you can also contact me on WhatsApp.",
            "warning",
        )

    return redirect(url_for("home", _anchor="contact"))


@app.post("/api/chat-stream")
def chat_stream() -> Response:
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        message = safe_text(request.form.get("message"), 4000)
        try:
            raw_history = json.loads(request.form.get("history", "[]"))
        except json.JSONDecodeError:
            raw_history = []
        attachments, attachment_error = read_chat_attachments()
        if attachment_error:
            return Response(json.dumps({"error": attachment_error}), status=400, content_type="application/json")
    else:
        body = request.get_json(silent=True) or {}
        message = safe_text(body.get("message"), 4000)
        raw_history = body.get("history")
        attachments = []
    history = sanitise_history(raw_history)

    if not message and not attachments:
        return Response(
            json.dumps({"error": "Message is required."}),
            status=400,
            content_type="application/json",
        )

    def send_event(payload: dict[str, Any]) -> str:
        return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

    @stream_with_context
    def generate():
        threading.Thread(
            target=forward_chat_copy_safely,
            args=(message, attachments),
            daemon=True,
        ).start()

        if openai_client is None:
            fallback_message = message or "The visitor shared a project attachment."
            yield send_event({"type": "delta", "text": local_assistant_reply(fallback_message)})
            yield send_event({"type": "done"})
            return

        user_content: list[dict[str, Any]] = [
            {"type": "input_text", "text": message or "Please review the attached project file."}
        ]
        for item in attachments:
            encoded = base64.b64encode(item["data"]).decode("ascii")
            data_url = f"data:{item['content_type']};base64,{encoded}"
            if item["content_type"].startswith("image/"):
                user_content.append({"type": "input_image", "image_url": data_url})
            else:
                user_content.append({"type": "input_file", "filename": item["name"], "file_data": data_url})
        conversation = [*history, {"role": "user", "content": user_content}]

        try:
            stream = openai_client.responses.create(
                model=OPENAI_MODEL,
                instructions=(
                    "You are Shubham Chauhan's professional portfolio AI assistant. "
                    "Shubham provides Python, Flask, FastAPI, Flutter, Android, iOS, "
                    "ERP, e-commerce, business website, desktop software, API and admin "
                    "dashboard development. Understand abbreviations and spelling errors. "
                    "Carefully analyse attached images and PDFs when present. Help the visitor "
                    "clarify requirements and recommend relevant modules using only supplied "
                    "information and the capabilities listed here. Explicitly say when information "
                    "is uncertain or needs human confirmation. "
                    "Never invent verified clients, fixed pricing, guaranteed timelines, "
                    "credentials or capabilities not stated here. Ask no more than one "
                    "useful follow-up question per response. Do not add a signature because the "
                    "application adds it automatically. Keep replies professional, clear, friendly "
                    "and normally below 220 words."
                ),
                input=conversation,
                stream=True,
            )

            emitted = False
            for event in stream:
                event_type = getattr(event, "type", "")
                if event_type == "response.output_text.delta":
                    delta = getattr(event, "delta", "")
                    if delta:
                        emitted = True
                        yield send_event({"type": "delta", "text": delta})
                elif event_type == "error":
                    raise RuntimeError(str(event))

            if not emitted:
                yield send_event(
                    {
                        "type": "delta",
                        "text": "I could not generate a reply. Please try again.",
                    }
                )
            yield send_event({"type": "done"})

        except Exception:
            app.logger.exception("OpenAI streaming request failed.")
            yield send_event(
                {
                    "type": "notice",
                    "text": "AI service is temporarily unavailable; showing a local assistant reply.",
                }
            )
            fallback_message = message or "The visitor shared a project attachment."
            yield send_event({"type": "delta", "text": local_assistant_reply(fallback_message)})
            yield send_event({"type": "done"})

    return Response(
        generate(),
        content_type="text/event-stream; charset=utf-8",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


def sync_request_authorized() -> bool:
    expected_values = {
        value for value in (
            os.getenv("SHEET_SYNC_SECRET", "").strip(),
            os.getenv("CRON_SECRET", "").strip(),
        ) if value
    }
    if not expected_values:
        return False
    supplied = request.headers.get("Authorization", "").removeprefix("Bearer ")
    supplied = supplied or request.headers.get("X-Sync-Secret", "")
    return bool(supplied) and any(
        secrets.compare_digest(supplied, expected) for expected in expected_values
    )


@app.route("/api/sync/google-sheet", methods=["GET", "POST"])
def sync_google_sheet_endpoint() -> tuple[dict[str, Any], int] | dict[str, Any]:
    if not sync_request_authorized():
        return {"status": "error", "message": "Unauthorized"}, 401
    try:
        result = sync_google_sheet_to_database()
        return {"status": "ok", **result}
    except Exception:
        app.logger.exception("Google Sheet reconciliation failed")
        return {"status": "error", "message": "Synchronization failed"}, 500


@app.get("/health")
def health() -> dict[str, Any]:
    return {
        "status": "ok",
        "ai_configured": bool(openai_client),
        "email_configured": bool(
            MAIL_USERNAME and MAIL_PASSWORD and MAIL_RECEIVER
        ),
        "google_sheets_configured": bool(
            GOOGLE_SHEET_ID
            and (
                GOOGLE_SERVICE_ACCOUNT_JSON_BASE64
                or GOOGLE_SERVICE_ACCOUNT_PATH.is_file()
            )
        ),
        "database_configured": enquiry_store.configured(),
    }


@app.after_request
def add_security_headers(response: Response) -> Response:
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    return response


if __name__ == "__main__":
    initialise_database()
    debug_enabled = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    app.run(host="127.0.0.1", port=5000, debug=debug_enabled, threaded=True)
